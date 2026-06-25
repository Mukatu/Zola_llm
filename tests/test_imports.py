"""Tests du framework Import/Export Excel (pur + round-trip endpoint SQLite)."""

from __future__ import annotations

from contextlib import asynccontextmanager
from io import BytesIO
from typing import Any

import openpyxl
from httpx import ASGITransport, AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from zolaos.api.dependencies import get_router_client
from zolaos.api.main import create_app
from zolaos.core.settings import Settings, get_settings
from zolaos.db.session import get_session
from zolaos.db.store_models import StoreBase
from zolaos.imports.framework import (
    build_pole_template,
    build_template,
    parse_pole,
    parse_sheet,
    validate_row,
)
from zolaos.imports.mapping import apply_mapping, propose_mapping
from zolaos.imports.registry import POLES, REGISTRY
from zolaos.llm.base import GenerationResult

_EMP_COLS = [
    "matricule",
    "nom_complet",
    "genre",
    "date_naissance",
    "date_embauche",
    "poste",
    "departement",
    "manager_matricule",
    "categorie",
    "code_emploi",
    "salaire_base_xaf",
    "quotite",
    "statut",
]


def _emp_xlsx(rows: list[dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employés"
    ws.append(_EMP_COLS)
    for r in rows:
        ws.append([r.get(c, "") for c in _EMP_COLS])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_validate_row() -> None:
    spec = REGISTRY["employees"]
    ok, errs = validate_row(
        spec, {"matricule": "E1", "nom_complet": "Awa", "date_embauche": "2024-01-01"}
    )
    assert errs == []
    assert ok is not None and ok["matricule"] == "E1"

    _, errs2 = validate_row(spec, {"nom_complet": "X"})  # matricule + date_embauche manquants
    assert any("matricule" in e for e in errs2)

    _, errs3 = validate_row(
        spec, {"matricule": "E1", "nom_complet": "X", "date_embauche": "2024-01-01", "genre": "Z"}
    )
    assert any("genre" in e for e in errs3)


def test_template_and_parse_roundtrip() -> None:
    data = build_template(REGISTRY["employees"])
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "Dictionnaire" in wb.sheetnames
    parsed = parse_sheet(
        _emp_xlsx([{"matricule": "E1", "nom_complet": "Awa", "date_embauche": "2024-01-01"}])
    )
    assert parsed[0]["matricule"] == "E1"


def _settings() -> Settings:
    return Settings(
        POSTGRES_PASSWORD_APP="x", POSTGRES_PASSWORD_MIGRATIONS="x", JWT_SECRET="x" * 32
    )


class _FakeLLM:
    """Client LLM factice : renvoie un JSON de mapping figé (aucun réseau)."""

    provider = "fake"

    def __init__(self, content: str = "{}") -> None:
        self._content = content

    async def generate(self, messages, *, model, options=None):  # type: ignore[no-untyped-def]
        return GenerationResult(content=self._content, model=model, provider="fake")

    async def stream(self, *a, **k):  # type: ignore[no-untyped-def]  # pragma: no cover
        yield ""

    async def health(self) -> bool:  # pragma: no cover
        return True


@asynccontextmanager
async def _client(tmp_path, llm: Any | None = None):  # type: ignore[no-untyped-def]
    engine = create_async_engine(f"sqlite+aiosqlite:///{tmp_path}/imp.db")
    async with engine.begin() as conn:
        await conn.run_sync(StoreBase.metadata.create_all)
    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async def _override():
        async with factory() as s:
            yield s

    settings = _settings()
    app = create_app(settings=settings)
    app.dependency_overrides[get_session] = _override
    app.dependency_overrides[get_settings] = lambda: settings
    app.dependency_overrides[get_router_client] = lambda: (llm or _FakeLLM())
    client = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
    try:
        yield client
    finally:
        await client.aclose()
        await engine.dispose()


async def test_import_dry_run_then_commit_idempotent(tmp_path) -> None:  # type: ignore[no-untyped-def]
    async with _client(tmp_path) as ac:
        # modèle téléchargeable
        r = await ac.get("/v1/erp/import/template/employees")
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]

        xlsx = _emp_xlsx(
            [
                {
                    "matricule": "E1",
                    "nom_complet": "Awa",
                    "date_embauche": "2024-01-01",
                    "salaire_base_xaf": "500000",
                },
                {"nom_complet": "SansMatricule"},  # rejetée
            ]
        )

        # dry-run : 1 valide, 1 erreur, rien d'importé
        r = await ac.post("/v1/erp/import/employees?dry_run=true", content=xlsx)
        body = r.json()
        assert body["valides"] == 1
        assert len(body["erreurs"]) == 1
        assert (await ac.get("/v1/erp/employees")).json()["employees"] == []

        # import réel
        r = await ac.post("/v1/erp/import/employees", content=xlsx)
        assert r.json()["importes"] == 1
        assert len((await ac.get("/v1/erp/employees")).json()["employees"]) == 1

        # ré-import → upsert (mise à jour, pas de doublon)
        r = await ac.post("/v1/erp/import/employees", content=xlsx)
        assert r.json()["mis_a_jour"] == 1
        assert len((await ac.get("/v1/erp/employees")).json()["employees"]) == 1


def test_pole_template_has_one_sheet_per_entity() -> None:
    pole = POLES["rh"]
    wb = openpyxl.load_workbook(BytesIO(build_pole_template(pole)))
    for spec in pole.entities:
        assert spec.label[:31] in wb.sheetnames
    assert "Dictionnaire" in wb.sheetnames


def _fill_pole_sheet(wb: Any, label: str, cols: list[str], rows: list[dict[str, Any]]) -> None:
    ws = wb[label[:31]]
    # remplace tout sauf l'en-tête généré (déjà présent)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])


async def test_import_pole_rh_dispatches_per_sheet(tmp_path) -> None:  # type: ignore[no-untyped-def]
    pole = POLES["rh"]
    async with _client(tmp_path) as ac:
        # classeur généré par le backend → on le remplit feuille par feuille
        r = await ac.get("/v1/erp/import/template/pole/rh")
        assert r.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(r.content))

        _fill_pole_sheet(
            wb,
            REGISTRY["employees"].label,
            [c.name for c in REGISTRY["employees"].columns],
            [{"matricule": "E1", "nom_complet": "Awa", "date_embauche": "2024-01-01"}],
        )
        _fill_pole_sheet(
            wb,
            REGISTRY["job_roles"].label,
            [c.name for c in REGISTRY["job_roles"].columns],
            [{"code_emploi": "DEV", "intitule": "Développeur", "activites": "coder;tester"}],
        )
        bio = BytesIO()
        wb.save(bio)
        classeur = bio.getvalue()

        # dry-run : rapport par entité
        r = await ac.post("/v1/erp/import/pole/rh?dry_run=true", content=classeur)
        rapport = r.json()["rapport"]
        assert rapport["employees"]["valides"] == 1
        assert rapport["job_roles"]["valides"] == 1

        # import réel
        r = await ac.post("/v1/erp/import/pole/rh", content=classeur)
        rapport = r.json()["rapport"]
        assert rapport["employees"]["importes"] == 1
        assert rapport["job_roles"]["importes"] == 1
        assert len((await ac.get("/v1/erp/employees")).json()["employees"]) == 1

    assert pole.pole == "rh"


def test_parse_pole_ignores_unknown_sheets() -> None:
    pole = POLES["compta"]
    wb = openpyxl.load_workbook(BytesIO(build_pole_template(pole)))
    parsed = parse_pole_bytes(wb, pole)
    assert "invoices" in parsed


def parse_pole_bytes(wb: Any, pole: Any) -> dict[str, Any]:
    bio = BytesIO()
    wb.save(bio)
    return parse_pole(bio.getvalue(), pole)


# --------------------------------------------------------------- IMP-3 : mapping


def _sheet_xlsx(title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.append(headers)
    for r in rows:
        ws.append(r)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_propose_mapping_aliases_fuzzy_and_unresolved() -> None:
    spec = REGISTRY["employees"]
    headers = ["Matricule RH", "Nom et Prénom", "Date d'embauche", "Couleur préférée"]
    res = propose_mapping(spec, headers)
    assert res.mapping["Nom et Prénom"] == "nom_complet"  # alias normalisé
    assert res.mapping["Date d'embauche"] == "date_embauche"  # fuzzy
    assert res.mapping["Matricule RH"] == "matricule"
    assert "Couleur préférée" in res.non_resolus  # rien d'assez proche
    # un champ n'est jamais affecté deux fois
    assert len(set(res.mapping.values())) == len(res.mapping)


def test_apply_mapping_renames_keys() -> None:
    out = apply_mapping({"Nom et Prénom": "Awa", "x": 1}, {"Nom et Prénom": "nom_complet"})
    assert out == {"nom_complet": "Awa", "x": 1}


async def test_import_auto_map_synonym_headers(tmp_path) -> None:  # type: ignore[no-untyped-def]
    # en-têtes synonymes (pas les noms canoniques) → auto_map les rapproche
    xlsx = _sheet_xlsx(
        "Employés",
        ["Matricule RH", "Nom et Prénom", "Date d'embauche"],
        [["E1", "Awa", "2024-01-01"]],
    )
    async with _client(tmp_path) as ac:
        r = await ac.post("/v1/erp/import/employees?dry_run=true", content=xlsx)
        body = r.json()
        assert body["valides"] == 1
        assert body["mapping"]["renommages"]["Nom et Prénom"] == "nom_complet"

        r = await ac.post("/v1/erp/import/employees", content=xlsx)
        assert r.json()["importes"] == 1
        emps = (await ac.get("/v1/erp/employees")).json()["employees"]
        assert emps[0]["nom_complet"] == "Awa"


async def test_import_auto_map_disabled_rejects_synonyms(tmp_path) -> None:  # type: ignore[no-untyped-def]
    xlsx = _sheet_xlsx("Employés", ["Matricule RH", "Nom et Prénom"], [["E1", "Awa"]])
    async with _client(tmp_path) as ac:
        r = await ac.post("/v1/erp/import/employees?dry_run=true&auto_map=false", content=xlsx)
        body = r.json()
        assert body["valides"] == 0  # colonnes non reconnues → ligne rejetée
        assert body["mapping"] is None


async def test_inspect_deterministic_and_llm(tmp_path) -> None:  # type: ignore[no-untyped-def]
    xlsx = _sheet_xlsx("Employés", ["Nom et Prénom", "Identifiant agent"], [["Awa", "E1"]])
    # déterministe : "Identifiant agent" non résolu, "Nom et Prénom" résolu
    async with _client(tmp_path) as ac:
        r = await ac.post("/v1/erp/import/employees/inspect", content=xlsx)
        body = r.json()
        assert body["mapping"]["Nom et Prénom"] == "nom_complet"
        assert "Identifiant agent" in body["non_resolus"]
        assert body["suggestions_llm"] == {}  # use_llm absent → pas d'appel

    # augmentation LLM : le client factice propose matricule pour l'en-tête non résolu
    llm = _FakeLLM('{"Identifiant agent": "matricule"}')
    async with _client(tmp_path, llm=llm) as ac:
        r = await ac.post("/v1/erp/import/employees/inspect?use_llm=true", content=xlsx)
        body = r.json()
        assert body["suggestions_llm"]["Identifiant agent"] == "matricule"
