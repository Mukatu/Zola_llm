"""Tests PAIE-3 — DAS 1 : agrégation annuelle (moteur) + endpoints état annuel/DAS1/export."""

from __future__ import annotations

from contextlib import asynccontextmanager
from decimal import Decimal
from io import BytesIO

import openpyxl
from httpx import ASGITransport, AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from zolaos.agents.erp.das1 import LignePaie, Salarie, construire_das1
from zolaos.api.main import create_app
from zolaos.core.settings import Settings
from zolaos.db.session import get_session
from zolaos.db.store_models import StoreBase


def _settings() -> Settings:
    return Settings(
        POSTGRES_PASSWORD_APP="x", POSTGRES_PASSWORD_MIGRATIONS="x", JWT_SECRET="x" * 32
    )


@asynccontextmanager
async def _client(tmp_path):  # type: ignore[no-untyped-def]
    engine = create_async_engine(f"sqlite+aiosqlite:///{tmp_path}/das1.db")
    async with engine.begin() as conn:
        await conn.run_sync(StoreBase.metadata.create_all)
    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async def _override():
        async with factory() as s:
            yield s

    app = create_app(settings=_settings())
    app.dependency_overrides[get_session] = _override
    client = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
    try:
        yield client
    finally:
        await client.aclose()
        await engine.dispose()


def test_construire_das1_agrege() -> None:
    # cotisations salariales (retraite 4 %) déjà calculées au mois : 4 % de 500k = 20 000
    lignes = [
        LignePaie(
            matricule="E1",
            mois=1,
            brut_xaf="500000",
            cotisations_salariales_xaf="20000",
            irpp_xaf="20000",
        ),
        LignePaie(
            matricule="E1",
            mois=2,
            brut_xaf="500000",
            cotisations_salariales_xaf="20000",
            irpp_xaf="20000",
        ),
        LignePaie(
            matricule="E2",
            mois=1,
            brut_xaf="300000",
            cotisations_salariales_xaf="12000",
            irpp_xaf="5000",
        ),
    ]
    salaries = [Salarie(matricule="E1", nom="AWA", sexe="F", profession="Cadre")]
    das1 = construire_das1(lignes, salaries, exercice="2026")
    assert das1.nb_salaries == 2
    assert das1.total_brut_xaf == Decimal("1300000")  # 1M + 300k
    e1 = next(x for x in das1.etat_annuel if x.matricule == "E1")
    assert e1.mensuels_xaf[0] == Decimal("500000") and e1.mensuels_xaf[1] == Decimal("500000")
    assert e1.total_xaf == Decimal("1000000")
    d1 = next(x for x in das1.lignes if x.matricule == "E1")
    assert d1.nom == "AWA"
    # salaire plafonné = brut net de retraite : 1 000 000 − (2 × 20 000) = 960 000
    assert d1.salaire_plafonne_xaf == Decimal("960000")
    # base imposable = 80 % du salaire plafonné : 0,8 × 960 000 = 768 000
    assert d1.base_imposable_xaf == Decimal("768000")


async def test_das1_endpoints_and_export(tmp_path) -> None:  # type: ignore[no-untyped-def]
    async with _client(tmp_path) as ac:
        for mois in ("2026-01", "2026-02"):
            await ac.post(
                "/v1/erp/payslips",
                json={
                    "employee_matricule": "E1",
                    "periode": mois,
                    "brut_mensuel_xaf": "500000",
                    "allow_unvalidated": True,
                },
            )
        await ac.post(
            "/v1/erp/payslips",
            json={
                "employee_matricule": "E2",
                "periode": "2026-01",
                "brut_mensuel_xaf": "300000",
                "allow_unvalidated": True,
            },
        )

        etat = (await ac.get("/v1/erp/payroll/etat-annuel?annee=2026")).json()
        assert len(etat["mois"]) == 12
        e1 = next(x for x in etat["lignes"] if x["matricule"] == "E1")
        assert e1["total_xaf"] == "1000000"

        # identité DAS1 enrichie (PAIE-3c) via le registre du personnel
        await ac.post(
            "/v1/erp/employees",
            json={
                "matricule": "E1",
                "nom_complet": "AWA OKEMBA",
                "date_embauche": "2024-01-01",
                "situation_matrimoniale": "marie",
                "nationalite": "congolaise",
                "nb_enfants": 3,
                "livret_cnss": "LV-001",
            },
        )
        das1 = (await ac.get("/v1/erp/payroll/das1?annee=2026")).json()
        assert das1["nb_salaries"] == 2
        assert das1["totaux"]["brut_xaf"] == "1300000"
        assert "raison_sociale" in das1["employeur"]
        e1 = next(x for x in das1["lignes"] if x["matricule"] == "E1")
        assert e1["nom"] == "AWA OKEMBA"
        assert e1["situation_matrimoniale"] == "marie"
        assert e1["nb_enfants"] == 3
        assert e1["livret_cnss"] == "LV-001"

        r = await ac.get("/v1/erp/payroll/das1/export?annee=2026")
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        wb = openpyxl.load_workbook(BytesIO(r.content))
        assert {"ETAT ANNUEL BRUT & IRPP", "DAS 1"} <= set(wb.sheetnames)


# Points de référence relevés sur la DAS 1 réelle (CONGO TELECOM, exercice 2022) :
# (brut annuel, salaire plafonné = brut net retraite, base imposable = 80 % plafonné).
# Cf. règle légale : retraite CNSS 4 % plafonnée à 14 400 000/an, puis abattement 20 %.
_REF_DAS1_2022 = [
    (4550563, 4368541, 3494833),  # ABONDO
    (4100866, 3936833, 3149466),  # ABOUMBA
    (5085314, 4881905, 3905524),  # ADOUA
    (2932425, 2815130, 2252104),  # AKA
    (4617447, 4432749, 3546199),  # AKENZE
]
_TAUX_RETRAITE = Decimal("0.04")
_PLAFOND_RETRAITE_ANNUEL = Decimal("14400000")
_ABATTEMENT = Decimal("0.20")


def _q(v: Decimal) -> Decimal:
    return v.quantize(Decimal("1"), rounding="ROUND_HALF_UP")


def test_regle_das1_reproduit_le_fichier_reel() -> None:
    """La règle (retraite 4 % plafonnée puis base = 80 % du net) retombe sur les
    chiffres réels de la DAS 1 VBA. Le fichier arrondit la retraite *au mois* ; en
    repartant du brut annuel, l'écart résiduel reste du bruit d'arrondi (±10 XAF
    sur des montants de ~5 M, soit < 0,001 %), ce qui confirme l'identité de règle."""
    for brut, plafonne_ref, base_ref in _REF_DAS1_2022:
        b = Decimal(brut)
        retraite = _q(_TAUX_RETRAITE * min(b, _PLAFOND_RETRAITE_ANNUEL))
        plafonne = b - retraite
        base = _q((Decimal("1") - _ABATTEMENT) * plafonne)
        assert abs(plafonne - Decimal(plafonne_ref)) <= 10, (brut, plafonne, plafonne_ref)
        assert abs(base - Decimal(base_ref)) <= 10, (brut, base, base_ref)


async def test_das1_chaine_reelle_plafonne_et_base(tmp_path) -> None:  # type: ignore[no-untyped-def]
    """Bout en bout : émission de bulletins via le calculateur de paie → la DAS 1
    affiche salaire plafonné = brut net de retraite et base = 80 % du plafonné."""
    async with _client(tmp_path) as ac:
        # 12 mois à 400 000 : retraite 4 % = 16 000/mois (sous le plafond 1 200 000)
        for mois in range(1, 13):
            r = await ac.post(
                "/v1/erp/payslips",
                json={
                    "employee_matricule": "E1",
                    "periode": f"2026-{mois:02d}",
                    "brut_mensuel_xaf": "400000",
                    "allow_unvalidated": True,
                },
            )
            assert r.status_code == 201

        das1 = (await ac.get("/v1/erp/payroll/das1?annee=2026")).json()
        e1 = next(x for x in das1["lignes"] if x["matricule"] == "E1")
        # brut 4 800 000 − retraite 192 000 = 4 608 000 ; base = 80 % = 3 686 400
        assert e1["brut_annuel_xaf"] == "4800000"
        assert e1["salaire_plafonne_xaf"] == "4608000"
        assert e1["base_imposable_xaf"] == "3686400"


async def test_das1_rubriques_secondaires(tmp_path) -> None:  # type: ignore[no-untyped-def]
    """PAIE-3e : avantages en nature + indemnités non imposables déclarés sur les
    bulletins sont agrégés annuellement et exposés dans la DAS 1 + l'export."""
    async with _client(tmp_path) as ac:
        for mois in (1, 2):
            r = await ac.post(
                "/v1/erp/payslips",
                json={
                    "employee_matricule": "E9",
                    "periode": f"2026-{mois:02d}",
                    "brut_mensuel_xaf": "300000",
                    "avantages_nature_xaf": "25000",
                    "indemnites_non_imposables_xaf": "40000",
                    "allow_unvalidated": True,
                },
            )
            assert r.status_code == 201

        das1 = (await ac.get("/v1/erp/payroll/das1?annee=2026")).json()
        e9 = next(x for x in das1["lignes"] if x["matricule"] == "E9")
        assert e9["avantages_nature_xaf"] == "50000"  # 2 × 25 000
        assert e9["indemnites_non_imposables_xaf"] == "80000"  # 2 × 40 000
        # forfaits T.R. / TOL-CAMU non sourcés (barème par défaut) ⇒ 0
        assert e9["taxe_regionale_xaf"] == "0"
        assert e9["tol_camu_xaf"] == "0"
        assert das1["totaux"]["avantages_nature_xaf"] == "50000"
        assert das1["totaux"]["indemnites_non_imposables_xaf"] == "80000"

        r = await ac.get("/v1/erp/payroll/das1/export?annee=2026")
        wb = openpyxl.load_workbook(BytesIO(r.content))
        entetes = [c.value for c in wb["DAS 1"][6]]
        assert "(e) AVANTAGES EN NATURE" in entetes
        assert "(j) INDEMNITÉS NON IMPOS." in entetes
        assert "(i) T.R." in entetes
        assert "TOL / CAMU" in entetes


async def test_bareme_validation_leve_le_verrou(tmp_path) -> None:  # type: ignore[no-untyped-def]
    """PAIE-5 : le barème expose valeurs+sources et reste verrouillé tant qu'un
    expert ne l'a pas validé ; la validation autorise l'émission définitive."""
    async with _client(tmp_path) as ac:
        b = (await ac.get("/v1/erp/payroll/bareme")).json()
        assert b["valide_fichier"] is False
        assert b["effectivement_valide"] is False
        assert {"irpp", "its"} <= set(b["regimes"])
        assert len(b["sources"]) >= 1 and "confiance" in b["sources"][0]
        assert any(br["nom"] == "allocations_familiales" for br in b["cnss_branches"])

        # verrou actif : émission définitive (sans allow_unvalidated) refusée
        r = await ac.post(
            "/v1/erp/payslips",
            json={"employee_matricule": "X", "periode": "2026-01", "brut_mensuel_xaf": "300000"},
        )
        assert r.status_code == 409

        # validation experte → lève le verrou (audité)
        v = (
            await ac.post(
                "/v1/erp/payroll/bareme/validate",
                json={"validated": True, "validated_by": "DRH", "note": "Conforme LF 2026"},
            )
        ).json()
        assert v["effectivement_valide"] is True
        assert v["validation"]["validated_by"] == "DRH"
        assert v["validation"]["validated_at"] is not None

        r2 = await ac.post(
            "/v1/erp/payslips",
            json={"employee_matricule": "X", "periode": "2026-01", "brut_mensuel_xaf": "300000"},
        )
        assert r2.status_code == 201

        # révocation → le verrou se remet
        await ac.post("/v1/erp/payroll/bareme/validate", json={"validated": False})
        r3 = await ac.post(
            "/v1/erp/payslips",
            json={"employee_matricule": "Y", "periode": "2026-01", "brut_mensuel_xaf": "300000"},
        )
        assert r3.status_code == 409


async def test_bareme_editable_par_tenant(tmp_path) -> None:  # type: ignore[no-untyped-def]
    """PAIE-6a : éditer le barème (sans code) crée un override versionné, refait
    tomber la validation, et le nouveau taux s'applique à l'émission."""
    async with _client(tmp_path) as ac:
        b0 = (await ac.get("/v1/erp/payroll/bareme")).json()
        assert b0["source_donnees"] == "defaut"
        assert b0["abattement_irpp_taux"] == "0.20"

        # valider la graine
        await ac.post(
            "/v1/erp/payroll/bareme/validate",
            json={"validated": True, "validated_by": "DRH"},
        )

        # éditer l'abattement → 30 % (sans toucher au code)
        b1 = (
            await ac.put(
                "/v1/erp/payroll/bareme", json={"abattement_irpp_taux": "0.30", "edited_by": "DRH"}
            )
        ).json()
        assert b1["source_donnees"] == "tenant"
        assert b1["abattement_irpp_taux"] == "0.30"
        assert b1["version"].startswith("custom-")
        # nouvelle version ⇒ la validation est retombée
        assert b1["effectivement_valide"] is False

        # émission re-verrouillée tant que non re-validée
        r = await ac.post(
            "/v1/erp/payslips",
            json={"employee_matricule": "Z", "periode": "2026-01", "brut_mensuel_xaf": "300000"},
        )
        assert r.status_code == 409

        # re-valider la nouvelle version → émission OK + nouveau taux appliqué
        await ac.post(
            "/v1/erp/payroll/bareme/validate",
            json={"validated": True, "validated_by": "DRH"},
        )
        r2 = await ac.post(
            "/v1/erp/payslips",
            json={"employee_matricule": "Z", "periode": "2026-01", "brut_mensuel_xaf": "300000"},
        )
        assert r2.status_code == 201
        # base = (300000 − 4% retraite) × (1 − 0,30) = 288000 × 0,70 = 201 600
        assert Decimal(r2.json()["base_imposable_xaf"]) == Decimal("201600")
