"""Connecteur fichiers CSV / Excel (V2.2 §2.4, §4.4 — livré en standard).

Ingestion de fichiers plats (CSV, XLSX) + watcher de dossier (polling portable
Windows/Linux). Lecture seule pour RH/factures/banque ; écriture possible pour
exporter des écritures comptables en CSV (`push_journal_entry`).

Config attendue :
    {
      "path": "/data/employes.xlsx",   # fichier source (lecture)
      "format": "xlsx",                 # optionnel, déduit de l'extension
      "sheet": "Feuil1",                # XLSX uniquement (optionnel)
      "encoding": "utf-8",              # CSV (optionnel)
      "delimiter": ",",                 # CSV (optionnel)
      "journal_output_path": "/data/ecritures.csv"  # pour push_journal_entry
    }
"""

from __future__ import annotations

import asyncio
import csv
from pathlib import Path
from typing import Any

from zolaos.connectors.base import (
    ConnectorConfigError,
    FinanceConnector,
    HRConnector,
    InvoiceConnector,
    JournalConnector,
)
from zolaos.connectors.models import BankTransaction, Employee, Invoice, JournalEntry
from zolaos.core.logging import get_logger

_log = get_logger("zolaos.connectors.csv_excel")


class CsvExcelConnector(HRConnector, InvoiceConnector, FinanceConnector, JournalConnector):
    """Connecteur fichiers plats CSV/XLSX."""

    name = "csv_excel"

    # -- lecture brute --------------------------------------------------------

    def _resolve_format(self) -> str:
        fmt = self.config.get("format")
        if fmt:
            return fmt.lower()
        suffix = Path(self.config.get("path", "")).suffix.lower()
        if suffix in {".xlsx", ".xlsm"}:
            return "xlsx"
        if suffix in {".csv", ".txt"}:
            return "csv"
        raise ConnectorConfigError(
            f"Format indéterminé pour {self.config.get('path')!r} ; préciser config['format']."
        )

    def _read_rows_sync(self, path: str | None = None) -> list[dict[str, Any]]:
        src = path or self.config.get("path")
        if not src:
            raise ConnectorConfigError("config['path'] obligatoire pour CsvExcelConnector.")
        p = Path(src)
        if not p.is_file():
            raise ConnectorConfigError(f"Fichier introuvable : {p}")
        fmt = self._resolve_format()
        if fmt == "csv":
            with p.open(encoding=self.config.get("encoding", "utf-8"), newline="") as f:
                reader = csv.DictReader(f, delimiter=self.config.get("delimiter", ","))
                return [dict(r) for r in reader]
        # xlsx
        from openpyxl import load_workbook

        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb[self.config["sheet"]] if self.config.get("sheet") else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
        return [dict(zip(headers, r, strict=False)) for r in rows[1:]]

    async def _read_rows(self, path: str | None = None) -> list[dict[str, Any]]:
        return await asyncio.to_thread(self._read_rows_sync, path)

    def _to_canonical(self, row: dict[str, Any]) -> dict[str, Any]:
        return self.mapping.apply(row) if self.mapping is not None else dict(row)

    # -- capacités lecture ----------------------------------------------------

    async def list_employees(self, **filters: Any) -> list[Employee]:
        async with self._instrument("list_employees"):
            rows = await self._read_rows()
            return [Employee(**self._to_canonical(r)) for r in rows]

    async def list_invoices(self, **filters: Any) -> list[Invoice]:
        async with self._instrument("list_invoices"):
            rows = await self._read_rows()
            return [Invoice(**self._to_canonical(r)) for r in rows]

    async def read_invoice(self, invoice_id: str) -> Invoice:
        async with self._instrument("read_invoice"):
            for inv in await self.list_invoices():
                if inv.id_externe == invoice_id:
                    return inv
            raise ConnectorConfigError(f"Facture {invoice_id!r} absente du fichier.")

    async def list_bank_transactions(self, **filters: Any) -> list[BankTransaction]:
        async with self._instrument("list_bank_transactions"):
            rows = await self._read_rows()
            return [BankTransaction(**self._to_canonical(r)) for r in rows]

    # -- capacité écriture (export CSV) --------------------------------------

    def _append_journal_sync(self, out_path: str, entry: JournalEntry) -> str:
        p = Path(out_path)
        new_file = not p.exists()
        with p.open("a", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            if new_file:
                w.writerow(
                    ["date_ecriture", "journal", "reference", "compte", "libelle", "debit_xaf", "credit_xaf"]
                )
            for ligne in entry.lignes:
                w.writerow([
                    entry.date_ecriture.isoformat(), entry.journal, entry.reference or "",
                    ligne.compte, ligne.libelle, str(ligne.debit_xaf), str(ligne.credit_xaf),
                ])
        return str(p)

    async def push_journal_entry(self, entry: JournalEntry) -> str:
        out = self.config.get("journal_output_path")
        if not out:
            raise ConnectorConfigError(
                "push_journal_entry exige config['journal_output_path'] (CSV de sortie)."
            )
        if not entry.est_equilibree():
            raise ConnectorConfigError("Écriture déséquilibrée (débit ≠ crédit), refus.")
        async with self._instrument("push_journal_entry"):
            return await asyncio.to_thread(self._append_journal_sync, out, entry)

    # -- watcher de dossier (polling portable) -------------------------------

    async def watch(self, *, interval_seconds: float = 5.0):  # type: ignore[no-untyped-def]
        """Génère les chemins de fichiers nouveaux/modifiés dans config['watch_dir'].

        Polling (pas d'inotify) pour portabilité Windows/Linux. À consommer via
        `async for path in connector.watch(): ...`.
        """
        watch_dir = self.config.get("watch_dir")
        if not watch_dir:
            raise ConnectorConfigError("watch() exige config['watch_dir'].")
        pattern = self.config.get("watch_glob", "*.csv")
        seen: dict[str, float] = {}
        while True:
            for f in Path(watch_dir).glob(pattern):
                mtime = f.stat().st_mtime
                if seen.get(str(f)) != mtime:
                    seen[str(f)] = mtime
                    yield str(f)
            await asyncio.sleep(interval_seconds)
