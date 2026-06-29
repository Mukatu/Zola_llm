"""Système de référence léger — Factures (CRUD) + clôture continue.

Profil box. Persistance des factures (`store_invoices`) + endpoint de
**réconciliation temps réel** : à chaque lot de mouvements, on relettre le
registre stocké (clôture continue). Multi-tenant via `tenant_id` (défaut local).
"""

from __future__ import annotations

from dataclasses import asdict
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Response, status
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession

from zolaos.agents.erp.achats import (
    OffreFournisseur,
    Supplier,
    comparer_offres,
    score_fournisseur,
    verifier_conformite,
)
from zolaos.agents.erp.compta import ChartOfAccounts, JournalValidator
from zolaos.agents.erp.das1 import Das1, LignePaie, Salarie, construire_das1, libelle_mois
from zolaos.agents.erp.engagements import (
    Engagement,
    PilotageBudgetaire,
    detect_alertes,
    engagement_stats,
    pilotage_budgetaire,
)
from zolaos.agents.erp.facility import Asset, Echeance, echeances_dues, maintenances_dues
from zolaos.agents.erp.hse import (
    Incident,
    Risque,
    cartographie_risques,
    statistiques_incidents,
    taux_frequence,
    taux_gravite,
)
from zolaos.agents.erp.inventory import (
    SEUIL_VALIDATION_DEFAUT_XAF,
    ArticleStock,
    PilotageStock,
    StockInsuffisant,
    appliquer_mouvement,
    estimer_valeur_mouvement,
    pilotage_stock,
    requiert_double_validation,
)
from zolaos.agents.erp.payroll import (
    PayrollCalculator,
    PayrollScaleNotValidated,
    load_payroll_scale,
)
from zolaos.agents.erp.reconciliation import reconcilier
from zolaos.agents.erp.supply import StockItem, alertes_rupture, analyser_reappro
from zolaos.agents.erp.treasury import (
    SEUIL_DECAISSEMENT_DEFAUT_XAF,
    CompteTresorerie,
    FluxPrevu,
    FluxRapprochable,
    FluxTresorerie,
    IndicateursTreso,
    LigneReleve,
    Previsionnel,
    indicateurs_tresorerie,
    position_tresorerie,
    previsionnel_tresorerie,
    rapprocher,
)
from zolaos.api.v1.config import get_config_service
from zolaos.connectors.models import BankTransaction, Invoice, JournalEntry, JournalLine
from zolaos.core.personalization import TenantConfigService
from zolaos.db.session import get_session
from zolaos.db.store_repo import (
    AssetRepository,
    BankAccountRepository,
    CashFlowRepository,
    EcheanceRepository,
    EmployeeRepository,
    EngagementRepository,
    IncidentRepository,
    InvoiceRepository,
    JournalRepository,
    PayslipRepository,
    PurchaseBudgetRepository,
    PurchaseOrderRepository,
    RisqueRepository,
    StockMoveRepository,
    StockRepository,
    SupplierRepository,
)

router = APIRouter(prefix="/v1/erp", tags=["store"])


# ---------------------------------------------------------------- schémas


class InvoiceIn(BaseModel):
    numero: str
    sens: str = "vente"
    tiers: str
    date_emission: date
    date_echeance: date | None = None
    montant_ht_xaf: Decimal = Decimal("0")
    montant_tva_xaf: Decimal | None = None
    montant_ttc_xaf: Decimal = Decimal("0")
    devise: str = "XAF"
    payee: bool = False
    country: str = "cg"


class InvoicePatch(BaseModel):
    tiers: str | None = None
    date_echeance: date | None = None
    montant_ttc_xaf: Decimal | None = None
    payee: bool | None = None


class ReconcileIn(BaseModel):
    transactions: list[BankTransaction] = Field(default_factory=list)
    fenetre_jours: int = Field(default=5, ge=0, le=60)


# ---------------------------------------------------------------- CRUD factures


@router.post("/invoices", status_code=status.HTTP_201_CREATED, summary="Créer une facture")
async def create_invoice(
    body: InvoiceIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await InvoiceRepository(session).create({**body.model_dump(), "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/invoices", summary="Lister les factures")
async def list_invoices(
    tenant_id: str = "local",
    sens: str | None = None,
    payee: bool | None = None,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await InvoiceRepository(session).list(tenant_id=tenant_id, sens=sens, payee=payee)
    return {"invoices": [r.to_dict() for r in rows]}


@router.get("/invoices/{invoice_id}", summary="Lire une facture")
async def get_invoice(
    invoice_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await InvoiceRepository(session).get(invoice_id, tenant_id=tenant_id)
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="invoice_not_found")
    return rec.to_dict()


@router.patch("/invoices/{invoice_id}", summary="Mettre à jour une facture")
async def patch_invoice(
    invoice_id: str,
    body: InvoicePatch,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await InvoiceRepository(session).update(
        invoice_id, tenant_id=tenant_id, fields=body.model_dump(exclude_none=True)
    )
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="invoice_not_found")
    await session.commit()
    return rec.to_dict()


@router.post("/invoices/{invoice_id}/pay", summary="Marquer payée")
async def pay_invoice(
    invoice_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await InvoiceRepository(session).mark_paid(invoice_id, tenant_id=tenant_id)
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="invoice_not_found")
    await session.commit()
    return rec.to_dict()


@router.delete("/invoices/{invoice_id}", summary="Supprimer une facture")
async def delete_invoice(
    invoice_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    ok = await InvoiceRepository(session).delete(invoice_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="invoice_not_found")
    await session.commit()
    return {"deleted": invoice_id}


# ---------------------------------------------------------------- clôture continue


@router.post("/reconcile", summary="Réconciliation temps réel (clôture continue)")
async def reconcile(
    body: ReconcileIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await InvoiceRepository(session).list(tenant_id=tenant_id, sens="vente", payee=False)
    invoices = [
        Invoice(
            id_externe=r.id,
            numero=r.numero,
            sens="vente",
            tiers=r.tiers,
            date_emission=r.date_emission,
            montant_ht_xaf=r.montant_ht_xaf,
            montant_ttc_xaf=r.montant_ttc_xaf,
            payee=r.payee,
            country=r.country,
        )
        for r in rows
    ]
    report = reconcilier(invoices, body.transactions, fenetre_jours=body.fenetre_jours)
    return {
        "rapprochements": [asdict(x) for x in report.rapprochements],
        "factures_en_attente": [asdict(x) for x in report.factures_en_attente],
        "mouvements_non_rapproches": report.mouvements_non_rapproches,
        "cloture": asdict(report.cloture) if report.cloture else None,
    }


# ---------------------------------------------------------------- écritures comptables


class JournalLineIn(BaseModel):
    compte: str
    libelle: str
    debit_xaf: Decimal = Decimal("0")
    credit_xaf: Decimal = Decimal("0")


class JournalEntryIn(BaseModel):
    date_ecriture: date
    journal: str = "OD"
    libelle: str
    reference: str | None = None
    lignes: list[JournalLineIn] = Field(min_length=1)
    country: str = "cg"
    allow_unbalanced: bool = False


@router.post("/journal", status_code=status.HTTP_201_CREATED, summary="Enregistrer une écriture")
async def create_entry(
    body: JournalEntryIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    entry = JournalEntry(
        date_ecriture=body.date_ecriture,
        journal=body.journal,
        libelle=body.libelle,
        reference=body.reference,
        country=body.country,
        lignes=[JournalLine(**ligne.model_dump()) for ligne in body.lignes],
    )
    report = JournalValidator(ChartOfAccounts.load(body.country)).validate(entry)
    if not report.ok and not body.allow_unbalanced:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={"code": "ecriture_invalide", "errors": report.errors},
        )
    lignes_json = [
        {
            "compte": ligne.compte,
            "libelle": ligne.libelle,
            "debit_xaf": str(ligne.debit_xaf),
            "credit_xaf": str(ligne.credit_xaf),
        }
        for ligne in body.lignes
    ]
    rec = await JournalRepository(session).create(
        {
            "tenant_id": tenant_id,
            "date_ecriture": body.date_ecriture,
            "journal": body.journal,
            "libelle": body.libelle,
            "reference": body.reference,
            "lignes": lignes_json,
            "total_debit_xaf": report.total_debit_xaf,
            "total_credit_xaf": report.total_credit_xaf,
            "equilibre": report.ok,
            "country": body.country,
        }
    )
    await session.commit()
    return {**rec.to_dict(), "validation": asdict(report)}


@router.get("/journal", summary="Lister les écritures")
async def list_entries(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await JournalRepository(session).list(tenant_id=tenant_id)
    return {"entries": [r.to_dict() for r in rows]}


@router.get("/journal/balance", summary="Balance vivante des comptes (clôture continue)")
async def trial_balance(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await JournalRepository(session).list(tenant_id=tenant_id)
    agg: dict[str, dict[str, Decimal]] = {}
    for entry in rows:
        for ligne in entry.lignes:
            acc = agg.setdefault(ligne["compte"], {"debit": Decimal("0"), "credit": Decimal("0")})
            acc["debit"] += Decimal(ligne["debit_xaf"])
            acc["credit"] += Decimal(ligne["credit_xaf"])
    comptes = [
        {
            "compte": compte,
            "debit_xaf": str(v["debit"]),
            "credit_xaf": str(v["credit"]),
            "solde_xaf": str(v["debit"] - v["credit"]),
        }
        for compte, v in sorted(agg.items())
    ]
    total_debit = sum((v["debit"] for v in agg.values()), Decimal("0"))
    total_credit = sum((v["credit"] for v in agg.values()), Decimal("0"))
    return {
        "comptes": comptes,
        "total_debit_xaf": str(total_debit),
        "total_credit_xaf": str(total_credit),
        "equilibre": total_debit == total_credit,
    }


@router.delete("/journal/{entry_id}", summary="Supprimer une écriture")
async def delete_entry(
    entry_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    ok = await JournalRepository(session).delete(entry_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="entry_not_found")
    await session.commit()
    return {"deleted": entry_id}


# ---------------------------------------------------------------- stocks


class StockItemIn(BaseModel):
    sku: str
    libelle: str
    quantite_actuelle: Decimal = Decimal("0")
    unite: str = "unité"
    conso_moyenne_jour: Decimal = Decimal("0")
    delai_appro_jours: int = 0
    stock_securite: Decimal = Decimal("0")
    country: str = "cg"


class StockPatch(BaseModel):
    quantite_actuelle: Decimal | None = None
    conso_moyenne_jour: Decimal | None = None
    delai_appro_jours: int | None = None
    stock_securite: Decimal | None = None


@router.post("/stock", status_code=status.HTTP_201_CREATED, summary="Créer un article de stock")
async def create_stock(
    body: StockItemIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await StockRepository(session).create({**body.model_dump(), "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/stock", summary="Lister les articles de stock")
async def list_stock(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await StockRepository(session).list(tenant_id=tenant_id)
    return {"items": [r.to_dict() for r in rows]}


@router.patch("/stock/{item_id}", summary="Mettre à jour un article")
async def patch_stock(
    item_id: str,
    body: StockPatch,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await StockRepository(session).update(
        item_id, tenant_id=tenant_id, fields=body.model_dump(exclude_none=True)
    )
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="item_not_found")
    await session.commit()
    return rec.to_dict()


@router.delete("/stock/{item_id}", summary="Supprimer un article")
async def delete_stock(
    item_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    ok = await StockRepository(session).delete(item_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="item_not_found")
    await session.commit()
    return {"deleted": item_id}


@router.post("/stock/analyze", summary="Réappro + alertes rupture sur le stock stocké")
async def analyze_stock(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await StockRepository(session).list(tenant_id=tenant_id)
    items = [
        StockItem(
            sku=r.sku,
            libelle=r.libelle,
            quantite_actuelle=r.quantite_actuelle,
            unite=r.unite,
            conso_moyenne_jour=r.conso_moyenne_jour,
            delai_appro_jours=r.delai_appro_jours,
            stock_securite=r.stock_securite,
            country=r.country,
        )
        for r in rows
    ]
    return {
        "suggestions": [asdict(s) for s in analyser_reappro(items)],
        "alertes": [asdict(a) for a in alertes_rupture(items, horizon_jours=30)],
    }


# ---------------------------------------------------------------- mouvements de stock (STOCK-1)


class StockMoveIn(BaseModel):
    reference: str
    type: str = "entree"  # entree | sortie | ajustement | transfert
    sku: str
    quantite: Decimal = Decimal("0")
    cout_unitaire_xaf: Decimal | None = None
    emplacement: str | None = None
    emplacement_dest: str | None = None
    lot: str | None = None
    date_peremption: date | None = None
    motif: str = ""
    date_mouvement: date
    country: str = "cg"


@router.post(
    "/stock-moves", status_code=status.HTTP_201_CREATED, summary="Créer un mouvement (brouillon)"
)
async def create_stock_move(
    body: StockMoveIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await StockMoveRepository(session).create({**body.model_dump(), "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/stock-moves", summary="Lister les mouvements de stock")
async def list_stock_moves(
    tenant_id: str = "local",
    sku: str | None = None,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await StockMoveRepository(session).list(tenant_id=tenant_id, sku=sku)
    return {"moves": [r.to_dict() for r in rows]}


@router.post(
    "/stock-moves/{move_id}/validate",
    summary="Valider un mouvement (workflow à seuil : N1 puis N2 au-delà du seuil)",
)
async def validate_stock_move(
    move_id: str,
    tenant_id: str = "local",
    seuil_xaf: Decimal = SEUIL_VALIDATION_DEFAUT_XAF,
    autoriser_negatif: bool = False,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    moves = StockMoveRepository(session)
    move = await moves.get(move_id, tenant_id=tenant_id)
    if move is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="move_not_found")
    if move.statut == "valide":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="deja_valide")

    items = StockRepository(session)
    item = await items.get_by_sku(move.sku, tenant_id=tenant_id)
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="article_inconnu")

    estimee = estimer_valeur_mouvement(
        type=move.type,
        quantite=move.quantite,
        pmp_actuel=item.pmp_xaf,
        cout_unitaire=move.cout_unitaire_xaf,
    )
    # Gouvernance : 1er palier (N1) si au-dessus du seuil et encore en brouillon.
    if move.statut == "brouillon" and requiert_double_validation(estimee, seuil_xaf):
        move.statut = "valide_n1"
        move.valeur_xaf = estimee  # aperçu (non appliqué)
        await session.flush()
        await session.commit()
        return {"move": move.to_dict(), "applique": False, "requiert_n2": True}

    # Application réelle : brouillon sous le seuil, ou N1 → N2 (2e validation).
    try:
        res = appliquer_mouvement(
            type=move.type,
            quantite=move.quantite,
            quantite_actuelle=item.quantite_actuelle,
            pmp_actuel=item.pmp_xaf,
            cout_unitaire=move.cout_unitaire_xaf,
            autoriser_negatif=autoriser_negatif,
        )
    except StockInsuffisant as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="stock_insuffisant"
        ) from exc

    item.quantite_actuelle = res.nouvelle_quantite
    item.pmp_xaf = res.nouveau_pmp_xaf
    move.valeur_xaf = res.valeur_mouvement_xaf
    move.statut = "valide"
    await session.flush()
    await session.commit()
    return {"move": move.to_dict(), "article": item.to_dict(), "applique": True}


@router.delete("/stock-moves/{move_id}", summary="Supprimer un mouvement (brouillon uniquement)")
async def delete_stock_move(
    move_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    moves = StockMoveRepository(session)
    move = await moves.get(move_id, tenant_id=tenant_id)
    if move is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="move_not_found")
    if move.statut == "valide":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="mouvement_valide")
    await moves.delete(move_id, tenant_id=tenant_id)
    await session.commit()
    return {"deleted": move_id}


# ---------------------------------------------------------------- inventaire physique (STOCK-3)


class InventaireLigne(BaseModel):
    sku: str
    quantite_comptee: Decimal


class InventaireIn(BaseModel):
    comptages: list[InventaireLigne] = Field(default_factory=list)
    date_inventaire: date | None = None


@router.post("/stock/inventory", summary="Inventaire physique : écarts → ajustements (brouillon)")
async def stock_inventory(
    body: InventaireIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    """Compare le compté au théorique ; crée un **ajustement brouillon** par écart
    (à valider ensuite via le workflow). Ne modifie pas directement le stock."""
    items = StockRepository(session)
    moves = StockMoveRepository(session)
    jour = body.date_inventaire or date.today()
    resultats: list[dict[str, Any]] = []
    for c in body.comptages:
        item = await items.get_by_sku(c.sku, tenant_id=tenant_id)
        if item is None:
            resultats.append({"sku": c.sku, "erreur": "article_inconnu"})
            continue
        ecart = c.quantite_comptee - item.quantite_actuelle
        move_id = None
        if ecart != 0:
            rec = await moves.create(
                {
                    "tenant_id": tenant_id,
                    "reference": f"INV-{jour.isoformat()}-{c.sku}",
                    "type": "ajustement",
                    "sku": c.sku,
                    "quantite": ecart,
                    "motif": "Inventaire physique",
                    "date_mouvement": jour,
                }
            )
            move_id = rec.id
        resultats.append(
            {
                "sku": c.sku,
                "theorique": str(item.quantite_actuelle),
                "comptee": str(c.quantite_comptee),
                "ecart": str(ecart),
                "ajustement_id": move_id,
            }
        )
    await session.commit()
    nb_ecarts = sum(1 for r in resultats if r.get("ajustement_id"))
    return {"date_inventaire": jour.isoformat(), "nb_ecarts": nb_ecarts, "resultats": resultats}


@router.get("/stock/peremption", summary="Alertes de péremption (lots proches/expirés)")
async def stock_peremption(
    tenant_id: str = "local",
    horizon_jours: int = 30,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await StockMoveRepository(session).list(tenant_id=tenant_id)
    jour = date.today()
    alertes: list[dict[str, Any]] = []
    for m in rows:
        if m.type != "entree" or m.statut != "valide" or m.date_peremption is None:
            continue
        jours = (m.date_peremption - jour).days
        if jours <= horizon_jours:
            alertes.append(
                {
                    "sku": m.sku,
                    "lot": m.lot,
                    "quantite": str(m.quantite),
                    "date_peremption": m.date_peremption.isoformat(),
                    "jours_restants": jours,
                    "niveau": "expire" if jours < 0 else "proche",
                }
            )
    alertes.sort(key=lambda a: a["jours_restants"])
    return {"horizon_jours": horizon_jours, "alertes": alertes}


# ---------------------------------------------------------------- pilotage stock (STOCK-4)


async def _articles_stock(session: AsyncSession, tenant_id: str) -> list[ArticleStock]:
    rows = await StockRepository(session).list(tenant_id=tenant_id)
    return [
        ArticleStock(
            sku=r.sku,
            libelle=r.libelle,
            quantite_actuelle=r.quantite_actuelle,
            conso_moyenne_jour=r.conso_moyenne_jour,
            stock_securite=r.stock_securite,
            pmp_xaf=r.pmp_xaf,
        )
        for r in rows
    ]


def build_pilotage_stock_xlsx(p: PilotageStock) -> bytes:
    """Classeur de pilotage stock : synthèse + détail par article (ABC, couverture)."""
    wb = openpyxl.Workbook()
    syn = wb.active
    syn.title = "Synthèse"
    syn.append(["Pilotage des stocks"])
    syn.append(["Articles", p.nb_articles])
    syn.append(["Valorisation totale (XAF)", float(p.valorisation_totale_xaf)])
    syn.append(["Ruptures", p.nb_rupture])
    syn.append(["Sous stock de sécurité", p.nb_sous_securite])
    syn.append(["Taux de rupture (%)", float(p.taux_rupture_pct)])
    syn.append(
        [
            "Couverture moyenne (j)",
            float(p.couverture_moyenne_jours) if p.couverture_moyenne_jours is not None else "",
        ]
    )
    syn.append(["Articles dormants", p.dormant_nb])
    syn.append(["Valeur dormante (XAF)", float(p.dormant_valeur_xaf)])
    syn.append(
        [
            "ABC (A/B/C)",
            f"{p.repartition_abc['A']}/{p.repartition_abc['B']}/{p.repartition_abc['C']}",
        ]
    )

    det = wb.create_sheet("Par article")
    det.append(
        [
            "SKU",
            "Libellé",
            "Quantité",
            "Valeur stock XAF",
            "Couverture j",
            "Rotation/an",
            "Classe ABC",
        ]
    )
    for a in p.par_article:
        det.append(
            [
                a.sku,
                a.libelle,
                float(a.quantite),
                float(a.valeur_stock_xaf),
                float(a.couverture_jours) if a.couverture_jours is not None else "",
                float(a.rotation_annuelle) if a.rotation_annuelle is not None else "",
                a.classe_abc,
            ]
        )
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


@router.get("/stock/pilotage", summary="Pilotage stock : valorisation, rotation, rupture, ABC")
async def stock_pilotage(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    p = pilotage_stock(await _articles_stock(session, tenant_id))
    return {"pilotage": asdict(p)}


@router.get("/stock/pilotage/export", summary="Exporter le pilotage stock (.xlsx)")
async def export_stock_pilotage(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> Response:
    p = pilotage_stock(await _articles_stock(session, tenant_id))
    return Response(
        content=build_pilotage_stock_xlsx(p),
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": 'attachment; filename="pilotage_stock.xlsx"'},
    )


# ---------------------------------------------------------------- Achats (P2c)


class SupplierIn(BaseModel):
    id_externe: str
    nom: str
    secteur: str | None = None
    note_qualite: Decimal = Field(default=Decimal("0"), ge=0, le=5)
    delai_moyen_jours: int = Field(default=0, ge=0)
    documents_conformite: list[str] = Field(default_factory=list)
    actif: bool = True
    country: str = "cg"


class SupplierPatch(BaseModel):
    nom: str | None = None
    secteur: str | None = None
    note_qualite: Decimal | None = None
    delai_moyen_jours: int | None = None
    documents_conformite: list[str] | None = None
    actif: bool | None = None


class PurchaseOrderLineIn(BaseModel):
    libelle: str
    montant_ht_xaf: Decimal = Decimal("0")


class PurchaseOrderIn(BaseModel):
    id_externe: str
    numero: str
    fournisseur: str
    objet: str = ""
    date_emission: date
    statut: str = "brouillon"
    lignes: list[PurchaseOrderLineIn] = Field(default_factory=list)
    montant_ht_xaf: Decimal = Decimal("0")
    montant_ttc_xaf: Decimal = Decimal("0")
    delai_livraison_jours: int = Field(default=0, ge=0)
    country: str = "cg"


class PurchaseOrderPatch(BaseModel):
    objet: str | None = None
    statut: str | None = None
    montant_ht_xaf: Decimal | None = None
    montant_ttc_xaf: Decimal | None = None
    delai_livraison_jours: int | None = None


def _supplier_of(rec: Any) -> Supplier:
    return Supplier(
        id_externe=rec.id_externe,
        nom=rec.nom,
        secteur=rec.secteur,
        note_qualite=rec.note_qualite,
        delai_moyen_jours=rec.delai_moyen_jours,
        documents_conformite=list(rec.documents_conformite or []),
        actif=rec.actif,
        country=rec.country,
    )


# ----- CRUD fournisseurs -----


@router.post("/suppliers", status_code=status.HTTP_201_CREATED, summary="Créer un fournisseur")
async def create_supplier(
    body: SupplierIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await SupplierRepository(session).create({**body.model_dump(), "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/suppliers", summary="Lister les fournisseurs")
async def list_suppliers(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await SupplierRepository(session).list(tenant_id=tenant_id)
    return {"suppliers": [r.to_dict() for r in rows]}


@router.get("/suppliers/scores", summary="Scoring + conformité des fournisseurs (sur le store)")
async def suppliers_scores(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await SupplierRepository(session).list(tenant_id=tenant_id)
    scores = []
    for r in rows:
        sup = _supplier_of(r)
        scores.append(
            {
                "id": r.id,
                **asdict(score_fournisseur(sup)),
                "conformite_manquante": verifier_conformite(sup),
            }
        )
    scores.sort(key=lambda s: s["score"], reverse=True)
    return {"scores": scores}


@router.patch("/suppliers/{supplier_id}", summary="Mettre à jour un fournisseur")
async def patch_supplier(
    supplier_id: str,
    body: SupplierPatch,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await SupplierRepository(session).update(
        supplier_id, tenant_id=tenant_id, fields=body.model_dump(exclude_none=True)
    )
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="supplier_not_found")
    await session.commit()
    return rec.to_dict()


@router.delete("/suppliers/{supplier_id}", summary="Supprimer un fournisseur")
async def delete_supplier(
    supplier_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    ok = await SupplierRepository(session).delete(supplier_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="supplier_not_found")
    await session.commit()
    return {"deleted": supplier_id}


# ----- CRUD bons de commande -----


@router.post(
    "/purchase-orders", status_code=status.HTTP_201_CREATED, summary="Créer un bon de commande"
)
async def create_po(
    body: PurchaseOrderIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    data = body.model_dump()
    data["lignes"] = [
        {"libelle": x["libelle"], "montant_ht_xaf": str(x["montant_ht_xaf"])}
        for x in data["lignes"]
    ]
    rec = await PurchaseOrderRepository(session).create({**data, "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/purchase-orders", summary="Lister les bons de commande")
async def list_pos(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await PurchaseOrderRepository(session).list(tenant_id=tenant_id)
    return {"purchase_orders": [r.to_dict() for r in rows]}


@router.get("/purchase-orders/compare", summary="Comparatif des BC (prix/délai, sur le store)")
async def compare_pos(
    tenant_id: str = "local",
    objet: str | None = None,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await PurchaseOrderRepository(session).list(tenant_id=tenant_id)
    offres = [
        OffreFournisseur(
            id_externe=r.id_externe,
            fournisseur=r.fournisseur,
            objet=r.objet,
            montant_ht_xaf=r.montant_ht_xaf,
            montant_ttc_xaf=r.montant_ttc_xaf,
            delai_livraison_jours=r.delai_livraison_jours,
            country=r.country,
        )
        for r in rows
        if objet is None or r.objet == objet
    ]
    return {"classement": [asdict(c) for c in comparer_offres(offres)]}


@router.patch("/purchase-orders/{po_id}", summary="Mettre à jour un bon de commande")
async def patch_po(
    po_id: str,
    body: PurchaseOrderPatch,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await PurchaseOrderRepository(session).update(
        po_id, tenant_id=tenant_id, fields=body.model_dump(exclude_none=True)
    )
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="po_not_found")
    await session.commit()
    return rec.to_dict()


@router.delete("/purchase-orders/{po_id}", summary="Supprimer un bon de commande")
async def delete_po(
    po_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    ok = await PurchaseOrderRepository(session).delete(po_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="po_not_found")
    await session.commit()
    return {"deleted": po_id}


class ReceiptLigneStock(BaseModel):
    sku: str
    quantite: Decimal = Decimal("0")
    cout_unitaire_xaf: Decimal | None = None


class ReceiptIn(BaseModel):
    # Lignes optionnelles : génèrent des entrées de stock (brouillon) à la réception.
    entrees: list[ReceiptLigneStock] = Field(default_factory=list)


@router.post(
    "/purchase-orders/{po_id}/receipt",
    summary="Réceptionner un BC → facture d'achat (+ entrées de stock optionnelles)",
)
async def receipt_po(
    po_id: str,
    body: ReceiptIn | None = None,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    pos = PurchaseOrderRepository(session)
    rec = await pos.get(po_id, tenant_id=tenant_id)
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="po_not_found")
    if rec.invoice_id is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="deja_receptionne")
    if rec.statut == "brouillon":
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="bon_non_emis")
    inv = await InvoiceRepository(session).create(
        {
            "tenant_id": tenant_id,
            "numero": rec.numero,
            "sens": "achat",
            "tiers": rec.fournisseur,
            "date_emission": rec.date_emission,
            "montant_ht_xaf": rec.montant_ht_xaf,
            "montant_ttc_xaf": rec.montant_ttc_xaf,
            "payee": False,
            "country": rec.country,
        }
    )
    await pos.update(
        po_id, tenant_id=tenant_id, fields={"invoice_id": inv.id, "statut": "receptionne"}
    )
    # Boucle Achats → Stock : entrées de stock (brouillon) à valider ensuite.
    entrees_stock: list[dict[str, Any]] = []
    if body is not None and body.entrees:
        moves = StockMoveRepository(session)
        for i, e in enumerate(body.entrees, start=1):
            mv = await moves.create(
                {
                    "tenant_id": tenant_id,
                    "reference": f"BC-{rec.numero}-{i}",
                    "type": "entree",
                    "sku": e.sku,
                    "quantite": e.quantite,
                    "cout_unitaire_xaf": e.cout_unitaire_xaf,
                    "motif": f"Réception BC {rec.numero}",
                    "date_mouvement": rec.date_emission,
                }
            )
            entrees_stock.append(mv.to_dict())
    await session.commit()
    return {
        "purchase_order": rec.to_dict(),
        "invoice": inv.to_dict(),
        "entrees_stock": entrees_stock,
    }


# ---------------------------------------------------------------- Engagements (Achats v2)


class EngagementIn(BaseModel):
    numero_eb: str
    numero_da: str | None = None
    numero_bc: str | None = None
    date_eb: date | None = None
    date_da: date | None = None
    date_bc: date | None = None
    direction: str | None = None
    service: str | None = None
    demandeur: str | None = None
    acheteur: str | None = None
    fournisseur: str | None = None
    description_besoin: str = ""
    description_da: str = ""
    description_bc: str = ""
    estimation_xaf: Decimal = Decimal("0")
    montant_xaf: Decimal = Decimal("0")
    statut_ebda: str = ""
    statut_bc: str = ""
    country: str = "cg"


class EngagementPatch(BaseModel):
    numero_da: str | None = None
    numero_bc: str | None = None
    date_da: date | None = None
    date_bc: date | None = None
    acheteur: str | None = None
    fournisseur: str | None = None
    description_da: str | None = None
    description_bc: str | None = None
    estimation_xaf: Decimal | None = None
    montant_xaf: Decimal | None = None
    statut_ebda: str | None = None
    statut_bc: str | None = None


def _engagement_of(rec: Any) -> Engagement:
    return Engagement(
        numero_eb=rec.numero_eb,
        numero_da=rec.numero_da,
        numero_bc=rec.numero_bc,
        date_eb=rec.date_eb,
        date_da=rec.date_da,
        date_bc=rec.date_bc,
        direction=rec.direction,
        service=rec.service,
        demandeur=rec.demandeur,
        acheteur=rec.acheteur,
        fournisseur=rec.fournisseur,
        estimation_xaf=rec.estimation_xaf,
        montant_xaf=rec.montant_xaf,
        statut_ebda=rec.statut_ebda,
        statut_bc=rec.statut_bc,
    )


@router.post("/engagements", status_code=status.HTTP_201_CREATED, summary="Créer un engagement")
async def create_engagement(
    body: EngagementIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await EngagementRepository(session).create({**body.model_dump(), "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/engagements", summary="Lister les engagements (chaîne EB→DA→BC)")
async def list_engagements(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await EngagementRepository(session).list(tenant_id=tenant_id)
    return {"engagements": [r.to_dict() for r in rows]}


@router.get(
    "/engagements/stats", summary="Indicateurs d'engagement (transformation, écarts, funnel)"
)
async def engagements_stats(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await EngagementRepository(session).list(tenant_id=tenant_id)
    engagements = [_engagement_of(r) for r in rows]
    stats = engagement_stats(engagements)
    alertes = detect_alertes(engagements)
    return {"stats": asdict(stats), "alertes": [asdict(a) for a in alertes]}


@router.patch("/engagements/{engagement_id}", summary="Mettre à jour un engagement")
async def patch_engagement(
    engagement_id: str,
    body: EngagementPatch,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await EngagementRepository(session).update(
        engagement_id, tenant_id=tenant_id, fields=body.model_dump(exclude_none=True)
    )
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="engagement_not_found")
    await session.commit()
    return rec.to_dict()


@router.delete("/engagements/{engagement_id}", summary="Supprimer un engagement")
async def delete_engagement(
    engagement_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    ok = await EngagementRepository(session).delete(engagement_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="engagement_not_found")
    await session.commit()
    return {"deleted": engagement_id}


# ---------------------------------------------------------------- Pilotage budgétaire (CDG)


class PurchaseBudgetIn(BaseModel):
    direction: str
    exercice: str
    budget_xaf: Decimal = Decimal("0")


@router.post(
    "/purchase-budgets", status_code=status.HTTP_201_CREATED, summary="Définir un budget achats"
)
async def set_purchase_budget(
    body: PurchaseBudgetIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    """Upsert : un seul budget par (direction, exercice)."""
    rec = await PurchaseBudgetRepository(session).upsert(
        tenant_id=tenant_id,
        direction=body.direction,
        exercice=body.exercice,
        budget_xaf=body.budget_xaf,
    )
    await session.commit()
    return rec.to_dict()


@router.get("/purchase-budgets", summary="Lister les budgets achats")
async def list_purchase_budgets(
    tenant_id: str = "local",
    exercice: str | None = None,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await PurchaseBudgetRepository(session).list(tenant_id=tenant_id)
    if exercice is not None:
        rows = [r for r in rows if r.exercice == exercice]
    return {"budgets": [r.to_dict() for r in rows]}


@router.delete("/purchase-budgets/{budget_id}", summary="Supprimer un budget achats")
async def delete_purchase_budget(
    budget_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    ok = await PurchaseBudgetRepository(session).delete(budget_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="budget_not_found")
    await session.commit()
    return {"deleted": budget_id}


_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def build_pilotage_xlsx(pilotage: PilotageBudgetaire, exercice: str | None) -> bytes:
    """Classeur CDG : synthèse + engagé/budget par direction + mensuel + fournisseurs."""
    wb = openpyxl.Workbook()

    syn = wb.active
    syn.title = "Synthèse"
    syn.append(["Pilotage des achats (engagé vs budget)"])
    syn.append(["Exercice", exercice or "tous"])
    syn.append([])
    syn.append(["Budget total (XAF)", float(pilotage.budget_total_xaf)])
    syn.append(["Engagé total (XAF)", float(pilotage.engage_total_xaf)])
    syn.append(["Reste (XAF)", float(pilotage.reste_total_xaf)])
    syn.append(["Consommation (%)", float(pilotage.consommation_pct)])

    dirs = wb.create_sheet("Par direction")
    dirs.append(
        ["Direction", "Budget XAF", "Engagé XAF", "Reste XAF", "Consommation %", "Niveau", "Nb"]
    )
    for d in pilotage.par_direction:
        dirs.append(
            [
                d.direction,
                float(d.budget_xaf),
                float(d.engage_xaf),
                float(d.reste_xaf),
                float(d.consommation_pct),
                d.niveau,
                d.nb,
            ]
        )

    mois = wb.create_sheet("Par mois")
    mois.append(["Mois", "Engagé XAF"])
    for s in pilotage.serie_mensuelle:
        mois.append([s.mois, float(s.engage_xaf)])

    four = wb.create_sheet("Top fournisseurs")
    four.append(["Fournisseur", "Engagé XAF", "Nb"])
    for f in pilotage.top_fournisseurs:
        four.append([f.fournisseur, float(f.engage_xaf), f.nb])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def _compute_pilotage(
    session: AsyncSession,
    *,
    tenant_id: str,
    exercice: str | None,
    date_debut: date | None,
    date_fin: date | None,
) -> PilotageBudgetaire:
    eng_rows = await EngagementRepository(session).list(tenant_id=tenant_id)
    budget_rows = await PurchaseBudgetRepository(session).list(tenant_id=tenant_id)

    def _ref(r: Any) -> date | None:
        return r.date_bc or r.date_da or r.date_eb

    def _garde(r: Any) -> bool:
        d = _ref(r)
        if exercice is not None and (d is None or d.strftime("%Y") != exercice):
            return False
        if date_debut is not None and (d is None or d < date_debut):
            return False
        if date_fin is not None and (d is None or d > date_fin):
            return False
        return True

    engagements = [_engagement_of(r) for r in eng_rows if _garde(r)]
    budget_par_direction: dict[str, Decimal] = {}
    for b in budget_rows:
        if exercice is None or b.exercice == exercice:
            budget_par_direction[b.direction] = (
                budget_par_direction.get(b.direction, Decimal("0")) + b.budget_xaf
            )
    return pilotage_budgetaire(engagements, budget_par_direction)


@router.get("/engagements/pilotage", summary="Pilotage CDG : engagé vs budget par direction")
async def engagements_pilotage(
    tenant_id: str = "local",
    exercice: str | None = None,
    date_debut: date | None = None,
    date_fin: date | None = None,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    pilotage = await _compute_pilotage(
        session, tenant_id=tenant_id, exercice=exercice, date_debut=date_debut, date_fin=date_fin
    )
    return {"exercice": exercice, "pilotage": asdict(pilotage)}


@router.get("/engagements/pilotage/export", summary="Exporter le pilotage CDG (.xlsx)")
async def export_pilotage(
    tenant_id: str = "local",
    exercice: str | None = None,
    date_debut: date | None = None,
    date_fin: date | None = None,
    session: AsyncSession = Depends(get_session),
) -> Response:
    pilotage = await _compute_pilotage(
        session, tenant_id=tenant_id, exercice=exercice, date_debut=date_debut, date_fin=date_fin
    )
    content = build_pilotage_xlsx(pilotage, exercice)
    nom = f"pilotage_achats_{exercice or 'tous'}"
    return Response(
        content=content,
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{nom}.xlsx"'},
    )


# ---------------------------------------------------------------- Trésorerie (TRESO-1)


class BankAccountIn(BaseModel):
    code: str
    libelle: str
    banque: str = ""
    type: str = "banque"  # banque | caisse | mobile_money
    devise: str = "XAF"
    iban: str | None = None
    solde_initial_xaf: Decimal = Decimal("0")
    country: str = "cg"


class BankAccountPatch(BaseModel):
    libelle: str | None = None
    banque: str | None = None
    type: str | None = None
    iban: str | None = None
    solde_initial_xaf: Decimal | None = None


class CashFlowIn(BaseModel):
    reference: str
    compte_code: str
    sens: str = "encaissement"  # encaissement | decaissement
    montant_xaf: Decimal = Decimal("0")
    date_operation: date
    date_prevue: date | None = None
    statut: str = "realise"  # prevu | realise
    categorie: str = ""
    tiers: str = ""
    libelle: str = ""
    mode: str = "virement"
    invoice_id: str | None = None
    country: str = "cg"


# ----- comptes -----


@router.post("/bank-accounts", status_code=status.HTTP_201_CREATED, summary="Créer un compte")
async def create_bank_account(
    body: BankAccountIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await BankAccountRepository(session).create({**body.model_dump(), "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/bank-accounts", summary="Lister les comptes de trésorerie")
async def list_bank_accounts(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await BankAccountRepository(session).list(tenant_id=tenant_id)
    return {"accounts": [r.to_dict() for r in rows]}


@router.patch("/bank-accounts/{account_id}", summary="Mettre à jour un compte")
async def patch_bank_account(
    account_id: str,
    body: BankAccountPatch,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await BankAccountRepository(session).update(
        account_id, tenant_id=tenant_id, fields=body.model_dump(exclude_none=True)
    )
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="account_not_found")
    await session.commit()
    return rec.to_dict()


@router.delete("/bank-accounts/{account_id}", summary="Supprimer un compte")
async def delete_bank_account(
    account_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    ok = await BankAccountRepository(session).delete(account_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="account_not_found")
    await session.commit()
    return {"deleted": account_id}


# ----- flux -----


@router.post("/cash-flows", status_code=status.HTTP_201_CREATED, summary="Créer un flux")
async def create_cash_flow(
    body: CashFlowIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await CashFlowRepository(session).create({**body.model_dump(), "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/cash-flows", summary="Lister les flux de trésorerie")
async def list_cash_flows(
    tenant_id: str = "local",
    compte_code: str | None = None,
    statut: str | None = None,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await CashFlowRepository(session).list(
        tenant_id=tenant_id, compte_code=compte_code, statut=statut
    )
    return {"flows": [r.to_dict() for r in rows]}


@router.delete("/cash-flows/{flow_id}", summary="Supprimer un flux")
async def delete_cash_flow(
    flow_id: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    ok = await CashFlowRepository(session).delete(flow_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="flow_not_found")
    await session.commit()
    return {"deleted": flow_id}


@router.get("/treasury/position", summary="Position de trésorerie (réalisée et projetée)")
async def treasury_position(
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    accounts = await BankAccountRepository(session).list(tenant_id=tenant_id)
    flows = await CashFlowRepository(session).list(tenant_id=tenant_id)
    comptes = [
        CompteTresorerie(
            code=a.code,
            libelle=a.libelle,
            type=a.type,
            devise=a.devise,
            solde_initial_xaf=a.solde_initial_xaf,
        )
        for a in accounts
    ]
    flux = [
        FluxTresorerie(
            compte_code=f.compte_code, sens=f.sens, montant_xaf=f.montant_xaf, statut=f.statut
        )
        for f in flows
    ]
    return {"position": asdict(position_tresorerie(comptes, flux))}


# ----- gouvernance : validation des décaissements (TRESO-3) -----


@router.post(
    "/cash-flows/{flow_id}/approve",
    summary="Approuver un décaissement (workflow à seuil : N1 puis N2)",
)
async def approve_cash_flow(
    flow_id: str,
    tenant_id: str = "local",
    seuil_xaf: Decimal = SEUIL_DECAISSEMENT_DEFAUT_XAF,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    flows = CashFlowRepository(session)
    flow = await flows.get(flow_id, tenant_id=tenant_id)
    if flow is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="flow_not_found")
    if flow.sens != "decaissement":
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="seuls_decaissements"
        )
    if flow.statut == "realise":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="deja_execute")

    # 1er palier (N1) si au-dessus du seuil et pas encore validé une fois.
    if flow.niveau_validation == "" and flow.montant_xaf > seuil_xaf:
        flow.niveau_validation = "n1"
        await session.flush()
        await session.commit()
        return {"flow": flow.to_dict(), "execute": False, "requiert_n2": True}

    # Exécution : sous le seuil, ou 2e validation (N2).
    flow.niveau_validation = "validee"
    flow.statut = "realise"
    await session.flush()
    await session.commit()
    return {"flow": flow.to_dict(), "execute": True}


# ----- rapprochement bancaire (TRESO-3) -----


class ReleveLigneIn(BaseModel):
    date: date
    montant_xaf: Decimal
    sens: str  # encaissement | decaissement
    libelle: str = ""


class ReconcileTresoIn(BaseModel):
    releve: list[ReleveLigneIn] = Field(default_factory=list)
    compte_code: str | None = None
    fenetre_jours: int = Field(default=5, ge=0, le=60)


@router.post("/treasury/reconcile", summary="Rapprochement bancaire (relevé vs flux réalisés)")
async def treasury_reconcile(
    body: ReconcileTresoIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    repo = CashFlowRepository(session)
    rows = await repo.list(tenant_id=tenant_id, compte_code=body.compte_code, statut="realise")
    candidats = [r for r in rows if not r.rapproche]
    flux = [
        FluxRapprochable(
            id=r.id,
            compte_code=r.compte_code,
            sens=r.sens,
            montant_xaf=r.montant_xaf,
            date_operation=r.date_operation,
        )
        for r in candidats
    ]
    releve = [
        LigneReleve(date=x.date, montant_xaf=x.montant_xaf, sens=x.sens, libelle=x.libelle)
        for x in body.releve
    ]
    res = rapprocher(flux, releve, fenetre_jours=body.fenetre_jours)
    # marque les flux appariés comme rapprochés
    appariés = {r.flux_id for r in res.rapprochements}
    for r in candidats:
        if r.id in appariés:
            r.rapproche = True
    await session.flush()
    await session.commit()
    return {
        "rapprochements": [asdict(r) for r in res.rapprochements],
        "flux_non_rapproches": res.flux_non_rapproches,
        "releve_non_rapproche": res.releve_non_rapproche,
        "taux_rapprochement_pct": str(res.taux_rapprochement_pct),
    }


# ----- pilotage : prévisionnel + indicateurs (TRESO-4) -----


async def _compute_treso_pilotage(
    session: AsyncSession, *, tenant_id: str, horizon_jours: int
) -> tuple[Previsionnel, IndicateursTreso]:
    accounts = await BankAccountRepository(session).list(tenant_id=tenant_id)
    flows = await CashFlowRepository(session).list(tenant_id=tenant_id)
    invoices = await InvoiceRepository(session).list(tenant_id=tenant_id)
    stock = await StockRepository(session).list(tenant_id=tenant_id)

    comptes = [
        CompteTresorerie(
            code=a.code,
            libelle=a.libelle,
            type=a.type,
            devise=a.devise,
            solde_initial_xaf=a.solde_initial_xaf,
        )
        for a in accounts
    ]
    flux_pos = [
        FluxTresorerie(
            compte_code=f.compte_code, sens=f.sens, montant_xaf=f.montant_xaf, statut=f.statut
        )
        for f in flows
    ]
    position = position_tresorerie(comptes, flux_pos)

    flux_prevus = [
        FluxPrevu(sens=f.sens, montant_xaf=f.montant_xaf, date=f.date_prevue or f.date_operation)
        for f in flows
        if f.statut == "prevu"
    ]
    prev = previsionnel_tresorerie(
        position.total_realise_xaf, flux_prevus, as_of=date.today(), horizon_jours=horizon_jours
    )

    encours_clients = sum(
        (i.montant_ttc_xaf for i in invoices if i.sens == "vente" and not i.payee), Decimal("0")
    )
    encours_fournisseurs = sum(
        (i.montant_ttc_xaf for i in invoices if i.sens == "achat" and not i.payee), Decimal("0")
    )
    ca = sum((i.montant_ttc_xaf for i in invoices if i.sens == "vente"), Decimal("0"))
    achats = sum((i.montant_ttc_xaf for i in invoices if i.sens == "achat"), Decimal("0"))
    valeur_stock = sum((s.quantite_actuelle * s.pmp_xaf for s in stock), Decimal("0"))

    mois = Decimal(horizon_jours) / Decimal("30")
    net_mensuel = (
        (prev.encaissements_total_xaf - prev.decaissements_total_xaf) / mois
        if mois > 0
        else Decimal("0")
    )
    indic = indicateurs_tresorerie(
        encours_clients=encours_clients,
        encours_fournisseurs=encours_fournisseurs,
        ca=ca,
        achats=achats,
        valeur_stock=valeur_stock,
        position_actuelle=position.total_realise_xaf,
        net_mensuel_prevu=net_mensuel,
    )
    return prev, indic


@router.get("/treasury/pilotage", summary="Pilotage : prévisionnel + DSO/DPO/BFR/runway")
async def treasury_pilotage(
    tenant_id: str = "local",
    horizon_jours: int = 90,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    prev, indic = await _compute_treso_pilotage(
        session, tenant_id=tenant_id, horizon_jours=horizon_jours
    )
    return {"previsionnel": asdict(prev), "indicateurs": asdict(indic)}


@router.get("/treasury/pilotage/export", summary="Exporter le pilotage trésorerie (.xlsx)")
async def export_treasury_pilotage(
    tenant_id: str = "local",
    horizon_jours: int = 90,
    session: AsyncSession = Depends(get_session),
) -> Response:
    prev, indic = await _compute_treso_pilotage(
        session, tenant_id=tenant_id, horizon_jours=horizon_jours
    )
    wb = openpyxl.Workbook()
    syn = wb.active
    syn.title = "Synthèse"
    syn.append(["Pilotage de trésorerie"])
    syn.append(["Position initiale (XAF)", float(prev.position_initiale_xaf)])
    syn.append(["Encaissements prévus (XAF)", float(prev.encaissements_total_xaf)])
    syn.append(["Décaissements prévus (XAF)", float(prev.decaissements_total_xaf)])
    syn.append(["Position finale projetée (XAF)", float(prev.position_finale_xaf)])
    syn.append(["Découvert prévu", prev.decouvert_periode or "—"])
    syn.append(["Encours clients (XAF)", float(indic.encours_clients_xaf)])
    syn.append(["Encours fournisseurs (XAF)", float(indic.encours_fournisseurs_xaf)])
    syn.append(["DSO (jours)", indic.dso_jours])
    syn.append(["DPO (jours)", indic.dpo_jours])
    syn.append(["BFR (XAF)", float(indic.bfr_xaf)])
    syn.append(["Runway (mois)", float(indic.runway_mois) if indic.runway_mois is not None else ""])

    det = wb.create_sheet("Prévisionnel")
    det.append(["Période", "Début", "Encaissements", "Décaissements", "Flux net", "Solde projeté"])
    for p in prev.periodes:
        det.append(
            [
                p.libelle,
                p.debut,
                float(p.encaissements_xaf),
                float(p.decaissements_xaf),
                float(p.flux_net_xaf),
                float(p.solde_projete_xaf),
            ]
        )
    bio = BytesIO()
    wb.save(bio)
    return Response(
        content=bio.getvalue(),
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": 'attachment; filename="pilotage_tresorerie.xlsx"'},
    )


# ---------------------------------------------------------------- Facility / Moyens généraux (OPS-1)


class AssetIn(BaseModel):
    id_externe: str
    libelle: str
    type_actif: str = "autre"
    maintenance_intervalle_jours: int = 0
    derniere_maintenance: date | None = None
    country: str = "cg"


class AssetPatch(BaseModel):
    libelle: str | None = None
    type_actif: str | None = None
    maintenance_intervalle_jours: int | None = None
    derniere_maintenance: date | None = None


class EcheanceIn(BaseModel):
    id_externe: str
    asset_id: str | None = None
    type_echeance: str = "autre"
    libelle: str
    date_echeance: date
    country: str = "cg"


@router.post("/assets", status_code=status.HTTP_201_CREATED, summary="Créer un actif")
async def create_asset(
    body: AssetIn, tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    rec = await AssetRepository(session).create({**body.model_dump(), "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/assets", summary="Lister les actifs")
async def list_assets(
    tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    rows = await AssetRepository(session).list(tenant_id=tenant_id)
    return {"assets": [r.to_dict() for r in rows]}


@router.patch("/assets/{asset_id}", summary="Mettre à jour un actif")
async def patch_asset(
    asset_id: str,
    body: AssetPatch,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await AssetRepository(session).update(
        asset_id, tenant_id=tenant_id, fields=body.model_dump(exclude_none=True)
    )
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="asset_not_found")
    await session.commit()
    return rec.to_dict()


@router.delete("/assets/{asset_id}", summary="Supprimer un actif")
async def delete_asset(
    asset_id: str, tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    ok = await AssetRepository(session).delete(asset_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="asset_not_found")
    await session.commit()
    return {"deleted": asset_id}


@router.post("/echeances", status_code=status.HTTP_201_CREATED, summary="Créer une échéance")
async def create_echeance(
    body: EcheanceIn, tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    rec = await EcheanceRepository(session).create({**body.model_dump(), "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/echeances", summary="Lister les échéances")
async def list_echeances(
    tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    rows = await EcheanceRepository(session).list(tenant_id=tenant_id)
    return {"echeances": [r.to_dict() for r in rows]}


@router.delete("/echeances/{echeance_id}", summary="Supprimer une échéance")
async def delete_echeance(
    echeance_id: str, tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    ok = await EcheanceRepository(session).delete(echeance_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="echeance_not_found")
    await session.commit()
    return {"deleted": echeance_id}


_TYPES_ACTIF = ("vehicule", "batiment", "equipement", "autre")
_TYPES_ECHEANCE = ("assurance", "controle", "contrat", "autre")


@router.get("/facility/echeancier", summary="Maintenances + échéances dues (sur le store)")
async def facility_echeancier_store(
    tenant_id: str = "local",
    horizon_jours: int = 30,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    asset_rows = await AssetRepository(session).list(tenant_id=tenant_id)
    ech_rows = await EcheanceRepository(session).list(tenant_id=tenant_id)
    assets = [
        Asset(
            id_externe=a.id_externe,
            libelle=a.libelle,
            type_actif=a.type_actif if a.type_actif in _TYPES_ACTIF else "autre",
            maintenance_intervalle_jours=a.maintenance_intervalle_jours,
            derniere_maintenance=a.derniere_maintenance,
            country=a.country,
        )
        for a in asset_rows
    ]
    echeances = [
        Echeance(
            id_externe=e.id_externe,
            asset_id=e.asset_id,
            type_echeance=e.type_echeance if e.type_echeance in _TYPES_ECHEANCE else "autre",
            libelle=e.libelle,
            date_echeance=e.date_echeance,
        )
        for e in ech_rows
    ]
    return {
        "maintenances": [asdict(m) for m in maintenances_dues(assets, horizon_jours=horizon_jours)],
        "echeances": [asdict(x) for x in echeances_dues(echeances, horizon_jours=horizon_jours)],
    }


# ---------------------------------------------------------------- HSE / RSE (OPS-1)


class RisqueIn(BaseModel):
    id_externe: str
    libelle: str
    probabilite: int = Field(default=1, ge=1, le=5)
    gravite: int = Field(default=1, ge=1, le=5)
    country: str = "cg"


class RisquePatch(BaseModel):
    libelle: str | None = None
    probabilite: int | None = None
    gravite: int | None = None


class IncidentIn(BaseModel):
    id_externe: str
    date_incident: date
    type_incident: str = "autre"
    gravite: str = "mineur"
    description: str = ""
    jours_arret: int = 0
    country: str = "cg"


_TYPES_INCIDENT = ("accident", "presqu_accident", "maladie", "environnement", "autre")
_GRAVITES_INCIDENT = ("mineur", "grave", "critique")


@router.post("/risques", status_code=status.HTTP_201_CREATED, summary="Créer un risque")
async def create_risque(
    body: RisqueIn, tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    rec = await RisqueRepository(session).create({**body.model_dump(), "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/risques", summary="Lister les risques")
async def list_risques(
    tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    rows = await RisqueRepository(session).list(tenant_id=tenant_id)
    return {"risques": [r.to_dict() for r in rows]}


@router.patch("/risques/{risque_id}", summary="Mettre à jour un risque")
async def patch_risque(
    risque_id: str,
    body: RisquePatch,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await RisqueRepository(session).update(
        risque_id, tenant_id=tenant_id, fields=body.model_dump(exclude_none=True)
    )
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="risque_not_found")
    await session.commit()
    return rec.to_dict()


@router.delete("/risques/{risque_id}", summary="Supprimer un risque")
async def delete_risque(
    risque_id: str, tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    ok = await RisqueRepository(session).delete(risque_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="risque_not_found")
    await session.commit()
    return {"deleted": risque_id}


@router.post("/incidents", status_code=status.HTTP_201_CREATED, summary="Déclarer un incident")
async def create_incident(
    body: IncidentIn, tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    rec = await IncidentRepository(session).create({**body.model_dump(), "tenant_id": tenant_id})
    await session.commit()
    return rec.to_dict()


@router.get("/incidents", summary="Lister les incidents")
async def list_incidents(
    tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    rows = await IncidentRepository(session).list(tenant_id=tenant_id)
    return {"incidents": [r.to_dict() for r in rows]}


@router.delete("/incidents/{incident_id}", summary="Supprimer un incident")
async def delete_incident(
    incident_id: str, tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    ok = await IncidentRepository(session).delete(incident_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="incident_not_found")
    await session.commit()
    return {"deleted": incident_id}


@router.get("/hse/cartographie", summary="Cartographie des risques (sur le store)")
async def hse_cartographie_store(
    tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    rows = await RisqueRepository(session).list(tenant_id=tenant_id)
    risques = [
        Risque(
            id_externe=r.id_externe,
            libelle=r.libelle,
            probabilite=r.probabilite,
            gravite=r.gravite,
            country=r.country,
        )
        for r in rows
    ]
    return {"risques": [asdict(x) for x in cartographie_risques(risques)]}


@router.get("/hse/indicators", summary="Indicateurs HSE (fréquence, gravité, incidents)")
async def hse_indicators_store(
    tenant_id: str = "local",
    heures_travaillees: int = 200000,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await IncidentRepository(session).list(tenant_id=tenant_id)
    incidents = [
        Incident(
            id_externe=i.id_externe,
            date_incident=i.date_incident,
            type_incident=i.type_incident if i.type_incident in _TYPES_INCIDENT else "autre",
            gravite=i.gravite if i.gravite in _GRAVITES_INCIDENT else "mineur",
            description=i.description,
            jours_arret=i.jours_arret,
            country=i.country,
        )
        for i in rows
    ]
    return {
        "statistiques": statistiques_incidents(incidents),
        "taux_frequence": str(taux_frequence(incidents, heures_travaillees=heures_travaillees)),
        "taux_gravite": str(taux_gravite(incidents, heures_travaillees=heures_travaillees)),
    }


# ---------------------------------------------------------------- Paie historisée (PAIE-1)

_payroll_calc = PayrollCalculator()


class PayslipIn(BaseModel):
    employee_matricule: str
    periode: str  # AAAA-MM
    brut_mensuel_xaf: Decimal = Field(..., ge=0)
    avantages_nature_xaf: Decimal = Field(default=Decimal("0"), ge=0)
    indemnites_non_imposables_xaf: Decimal = Field(default=Decimal("0"), ge=0)
    allow_unvalidated: bool = False
    country: str = "cg"


class PayslipPatch(BaseModel):
    statut: str | None = None
    date_paiement: date | None = None


@router.post(
    "/payslips", status_code=status.HTTP_201_CREATED, summary="Émettre un bulletin (historisé)"
)
async def create_payslip(
    body: PayslipIn,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    scale = load_payroll_scale(body.country)
    try:
        result = _payroll_calc.compute(
            body.brut_mensuel_xaf, scale=scale, allow_unvalidated=body.allow_unvalidated
        )
    except PayrollScaleNotValidated as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT, detail="bareme_non_valide"
        ) from exc
    rec = await PayslipRepository(session).upsert(
        {
            "tenant_id": tenant_id,
            "employee_matricule": body.employee_matricule,
            "periode": body.periode,
            "brut_xaf": result.brut_xaf,
            "cotisations_salariales": {k: str(v) for k, v in result.cotisations_salariales.items()},
            "total_cotisations_salariales_xaf": result.total_cotisations_salariales_xaf,
            "base_imposable_xaf": result.base_imposable_xaf,
            "irpp_xaf": result.irpp_xaf,
            "avantages_nature_xaf": body.avantages_nature_xaf,
            "indemnites_non_imposables_xaf": body.indemnites_non_imposables_xaf,
            "net_a_payer_xaf": result.net_a_payer_xaf,
            "cotisations_patronales": {k: str(v) for k, v in result.cotisations_patronales.items()},
            "cout_employeur_xaf": result.cout_employeur_xaf,
            "country": body.country,
        }
    )
    await session.commit()
    return rec.to_dict()


@router.get("/payslips", summary="Lister les bulletins (filtrable période/matricule)")
async def list_payslips(
    tenant_id: str = "local",
    periode: str | None = None,
    employee_matricule: str | None = None,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await PayslipRepository(session).list(
        tenant_id=tenant_id, periode=periode, employee_matricule=employee_matricule
    )
    return {"payslips": [r.to_dict() for r in rows]}


@router.patch("/payslips/{payslip_id}", summary="Valider / marquer payé un bulletin")
async def patch_payslip(
    payslip_id: str,
    body: PayslipPatch,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rec = await PayslipRepository(session).update(
        payslip_id, tenant_id=tenant_id, fields=body.model_dump(exclude_none=True)
    )
    if rec is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="payslip_not_found")
    await session.commit()
    return rec.to_dict()


@router.delete("/payslips/{payslip_id}", summary="Supprimer un bulletin")
async def delete_payslip(
    payslip_id: str, tenant_id: str = "local", session: AsyncSession = Depends(get_session)
) -> dict[str, Any]:
    ok = await PayslipRepository(session).delete(payslip_id, tenant_id=tenant_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="payslip_not_found")
    await session.commit()
    return {"deleted": payslip_id}


@router.get("/payroll/dashboard", summary="Pilotage masse salariale + déclaratif (par période)")
async def payroll_dashboard(
    tenant_id: str = "local",
    periode: str | None = None,
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    rows = await PayslipRepository(session).list(tenant_id=tenant_id, periode=periode)
    z = Decimal("0")
    brut = sum((r.brut_xaf for r in rows), z)
    net = sum((r.net_a_payer_xaf for r in rows), z)
    irpp = sum((r.irpp_xaf for r in rows), z)
    cot_sal = sum((r.total_cotisations_salariales_xaf for r in rows), z)
    cot_pat = sum(
        (sum((Decimal(v) for v in r.cotisations_patronales.values()), z) for r in rows), z
    )
    cout = sum((r.cout_employeur_xaf for r in rows), z)
    return {
        "periode": periode,
        "nb_bulletins": len(rows),
        "masse_salariale_brute_xaf": str(brut),
        "total_net_a_payer_xaf": str(net),
        "total_irpp_xaf": str(irpp),
        "total_cotisations_salariales_xaf": str(cot_sal),
        "total_cotisations_patronales_xaf": str(cot_pat),
        "cout_employeur_total_xaf": str(cout),
    }


# ---------------------------------------------------------------- DAS 1 (agrégation annuelle, PAIE-3)


def _employeur(config_service: TenantConfigService, tenant_id: str) -> dict[str, str]:
    """Infos du déclarant, lues dans la config du tenant (champs personnalisés)."""
    cfg = config_service.resolve("box", tenant_id=tenant_id)
    cp = cfg.champs_personnalises
    return {
        "raison_sociale": cp.get("employeur_raison_sociale", cfg.branding.nom_affichage),
        "matricule_cnss": cp.get("employeur_matricule_cnss", ""),
        "n_contribuable": cp.get("employeur_n_contribuable", ""),
        "bp": cp.get("employeur_bp", ""),
        "ville": cp.get("employeur_ville", ""),
    }


async def _build_das1(session: AsyncSession, *, tenant_id: str, annee: str) -> Das1:
    payslips = await PayslipRepository(session).list(tenant_id=tenant_id)
    lignes = [
        LignePaie(
            matricule=p.employee_matricule,
            mois=int(p.periode[5:7]),
            brut_xaf=p.brut_xaf,
            cotisations_salariales_xaf=p.total_cotisations_salariales_xaf,
            irpp_xaf=p.irpp_xaf,
            avantages_nature_xaf=p.avantages_nature_xaf,
            indemnites_non_imposables_xaf=p.indemnites_non_imposables_xaf,
        )
        for p in payslips
        if p.periode[:4] == annee and len(p.periode) >= 7
    ]
    employees = await EmployeeRepository(session).list(tenant_id=tenant_id)
    salaries = [
        Salarie(
            matricule=e.matricule,
            nom=e.nom_complet,
            sexe=e.genre,
            date_embauche=e.date_embauche,
            date_depart=e.date_sortie,
            profession=e.poste,
            livret_cnss=e.livret_cnss or "",
            n_contribuable=e.n_contribuable or "",
            situation_matrimoniale=e.situation_matrimoniale,
            nationalite=e.nationalite,
            nb_enfants=e.nb_enfants,
        )
        for e in employees
    ]
    scale = load_payroll_scale("cg")
    return construire_das1(
        lignes,
        salaries,
        exercice=annee,
        abattement_taux=scale.abattement_irpp_taux,
        taxe_regionale_xaf=scale.taxe_regionale_annuelle_xaf,
        tol_camu_xaf=scale.tol_camu_annuel_xaf,
    )


@router.get("/payroll/etat-annuel", summary="État annuel brut & IRPP (matrice salarié × 12 mois)")
async def payroll_etat_annuel(
    annee: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
) -> dict[str, Any]:
    das1 = await _build_das1(session, tenant_id=tenant_id, annee=annee)
    return {
        "exercice": annee,
        "mois": [libelle_mois(i) for i in range(1, 13)],
        "lignes": [
            {
                "matricule": e.matricule,
                "nom": e.nom,
                "mensuels_xaf": [str(v) for v in e.mensuels_xaf],
                "total_xaf": str(e.total_xaf),
                "irpp_annuel_xaf": str(e.irpp_annuel_xaf),
            }
            for e in das1.etat_annuel
        ],
        "total_brut_xaf": str(das1.total_brut_xaf),
        "total_irpp_xaf": str(das1.total_irpp_xaf),
    }


@router.get("/payroll/das1", summary="DAS 1 / CNSS 1 (par salarié) — exercice")
async def payroll_das1(
    annee: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
    config_service: TenantConfigService = Depends(get_config_service),
) -> dict[str, Any]:
    das1 = await _build_das1(session, tenant_id=tenant_id, annee=annee)
    return {
        "exercice": annee,
        "employeur": _employeur(config_service, tenant_id),
        "nb_salaries": das1.nb_salaries,
        "totaux": {
            "brut_xaf": str(das1.total_brut_xaf),
            "plafonne_xaf": str(das1.total_plafonne_xaf),
            "base_imposable_xaf": str(das1.total_base_imposable_xaf),
            "irpp_xaf": str(das1.total_irpp_xaf),
            "avantages_nature_xaf": str(das1.total_avantages_nature_xaf),
            "indemnites_non_imposables_xaf": str(das1.total_indemnites_non_imposables_xaf),
            "taxe_regionale_xaf": str(das1.total_taxe_regionale_xaf),
            "tol_camu_xaf": str(das1.total_tol_camu_xaf),
        },
        "lignes": [asdict(line) for line in das1.lignes],
    }


def build_das1_xlsx(das1: Das1, employeur: dict[str, str]) -> bytes:
    """Reproduit la DAS 1 (formulaire officiel) + l'état annuel sur 2 feuilles."""
    from openpyxl.styles import Alignment, Border, Font, Side

    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    title = Font(bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Feuille 1 : État annuel brut & IRPP
    e = wb.active
    e.title = "ETAT ANNUEL BRUT & IRPP"
    e.merge_cells("A1:O1")
    e["A1"] = (
        f"ÉTAT RÉCAPITULATIF DES SALAIRES BRUTS & IRPP — {employeur['raison_sociale']} — Exercice {das1.exercice}"
    )
    e["A1"].font = title
    e["A1"].alignment = center
    entete = ["N°", "NOMS", *[libelle_mois(i)[:4] for i in range(1, 13)], "TOTAL", "IRPP"]
    e.append([])
    e.append(entete)
    for c in e[3]:
        c.font = bold
        c.border = box
        c.alignment = center
    for i, ligne in enumerate(das1.etat_annuel, start=1):
        e.append(
            [
                i,
                ligne.nom,
                *[float(v) for v in ligne.mensuels_xaf],
                float(ligne.total_xaf),
                float(ligne.irpp_annuel_xaf),
            ]
        )
    e.append(
        ["", "TOTAL GÉNÉRAL", *[""] * 12, float(das1.total_brut_xaf), float(das1.total_irpp_xaf)]
    )
    for c in e[e.max_row]:
        c.font = bold

    # ---- Feuille 2 : DAS 1 (formulaire)
    d = wb.create_sheet("DAS 1")
    span = "A1:S1"
    d.merge_cells(span)
    d["A1"] = "RÉPUBLIQUE DU CONGO — DAS 1 / CNSS 1"
    d["A1"].font = title
    d["A1"].alignment = center
    d.merge_cells("A2:S2")
    d["A2"] = "DÉCLARATION ANNUELLE DES SALAIRES ET AUTRES RÉMUNÉRATIONS VERSÉES"
    d["A2"].alignment = center
    d.merge_cells("A3:S3")
    d["A3"] = (
        f"EMPLOYEUR : {employeur['raison_sociale']}   ·   MATRICULE CNSS : "
        f"{employeur['matricule_cnss']}   ·   N° CONTRIBUABLE : {employeur['n_contribuable']}"
    )
    d.merge_cells("A4:S4")
    d["A4"] = f"B.P : {employeur['bp']}   ·   {employeur['ville']}   ·   EXERCICE : {das1.exercice}"
    d.append([])
    cols = [
        "N°",
        "NOM - PRÉNOM",
        "SEXE",
        "SIT. MATRIM. (CMVD)",
        "NATIONALITÉ",
        "NB ENF.",
        "N° LIVRET CNSS",
        "N° CONTRIBUABLE",
        "PROFESSION",
        "EMBAUCHE",
        "DÉPART",
        "(f) SALAIRE BRUT",
        "SALAIRE PLAFONNÉ",
        "(e) AVANTAGES EN NATURE",
        "BASE IMPOSABLE (g=80%)",
        "(h) I.R.P.P.",
        "(i) T.R.",
        "TOL / CAMU",
        "(j) INDEMNITÉS NON IMPOS.",
    ]
    d.append(cols)
    for c in d[d.max_row]:
        c.font = bold
        c.border = box
        c.alignment = center
    for i, line in enumerate(das1.lignes, start=1):
        d.append(
            [
                i,
                line.nom,
                line.sexe,
                line.situation_matrimoniale,
                line.nationalite,
                line.nb_enfants,
                line.livret_cnss,
                line.n_contribuable,
                line.profession,
                line.date_embauche or "",
                line.date_depart or "",
                float(line.brut_annuel_xaf),
                float(line.salaire_plafonne_xaf),
                float(line.avantages_nature_xaf),
                float(line.base_imposable_xaf),
                float(line.irpp_xaf),
                float(line.taxe_regionale_xaf),
                float(line.tol_camu_xaf),
                float(line.indemnites_non_imposables_xaf),
            ]
        )
        for c in d[d.max_row]:
            c.border = box
    d.append(
        [
            "",
            "TOTAUX",
            *[""] * 9,
            float(das1.total_brut_xaf),
            float(das1.total_plafonne_xaf),
            float(das1.total_avantages_nature_xaf),
            float(das1.total_base_imposable_xaf),
            float(das1.total_irpp_xaf),
            float(das1.total_taxe_regionale_xaf),
            float(das1.total_tol_camu_xaf),
            float(das1.total_indemnites_non_imposables_xaf),
        ]
    )
    for c in d[d.max_row]:
        c.font = bold
    widths = (6, 24, 6, 16, 14, 8, 16, 16, 20, 12, 12, 16, 16, 18, 18, 14, 10, 12, 20)
    for col, w in zip("ABCDEFGHIJKLMNOPQRS", widths, strict=True):
        d.column_dimensions[col].width = w

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


@router.get("/payroll/das1/export", summary="Exporter la DAS 1 + état annuel (.xlsx, formulaire)")
async def export_das1(
    annee: str,
    tenant_id: str = "local",
    session: AsyncSession = Depends(get_session),
    config_service: TenantConfigService = Depends(get_config_service),
) -> Response:
    das1 = await _build_das1(session, tenant_id=tenant_id, annee=annee)
    content = build_das1_xlsx(das1, _employeur(config_service, tenant_id))
    return Response(
        content=content,
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="DAS1_{annee}.xlsx"'},
    )
