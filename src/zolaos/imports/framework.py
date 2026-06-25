"""Framework d'import/export piloté par schéma (déterministe).

Une `EntitySpec` déclare les colonnes attendues ; on en **génère** le template
`.xlsx`, on **valide** un upload (ligne par ligne) et on en extrait des
enregistrements prêts à persister. Aucune validation par LLM.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


@dataclass(frozen=True)
class Column:
    name: str
    kind: str = "str"  # str | int | decimal | date | bool | list
    required: bool = False
    enum: tuple[str, ...] | None = None
    help: str = ""
    aliases: tuple[str, ...] = field(default_factory=tuple)  # synonymes d'en-tête (IMP-3)


@dataclass(frozen=True)
class EntitySpec:
    entity: str
    label: str
    model: type[Any]
    columns: tuple[Column, ...]
    natural_key: tuple[str, ...] = field(default_factory=tuple)


def _coerce(col: Column, value: Any) -> tuple[Any, str | None]:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        if col.required:
            return None, f"« {col.name} » obligatoire"
        return None, None
    try:
        if col.kind == "str":
            out: Any = str(value).strip()
        elif col.kind == "int":
            out = int(value)
        elif col.kind == "decimal":
            out = Decimal(str(value))
        elif col.kind == "bool":
            out = str(value).strip().lower() in {"1", "true", "vrai", "oui", "yes", "x"}
        elif col.kind == "date":
            if isinstance(value, datetime):
                out = value.date()
            elif isinstance(value, date):
                out = value
            else:
                out = date.fromisoformat(str(value)[:10])
        elif col.kind == "list":
            if isinstance(value, (list | tuple)):
                out = list(value)
            else:
                out = [x.strip() for x in str(value).split(";") if x.strip()]
        else:
            out = value
    except (ValueError, InvalidOperation):
        return None, f"« {col.name} » : valeur invalide ({value!r})"
    if col.enum is not None and str(out) not in col.enum:
        return None, f"« {col.name} » : {out!r} hors {list(col.enum)}"
    return out, None


def validate_row(spec: EntitySpec, raw: dict[str, Any]) -> tuple[dict[str, Any] | None, list[str]]:
    """Coerce + valide une ligne → (enregistrement, erreurs)."""
    record: dict[str, Any] = {}
    errors: list[str] = []
    for col in spec.columns:
        value, err = _coerce(col, raw.get(col.name))
        if err is not None:
            errors.append(err)
        elif value is not None:
            record[col.name] = value
    for key in spec.natural_key:
        if key not in record:
            errors.append(f"clé « {key} » manquante")
    return (None if errors else record), errors


def build_template(spec: EntitySpec) -> bytes:
    """Génère le modèle .xlsx (données + dictionnaire + listes déroulantes)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = spec.label[:31]
    ws.append([c.name + ("*" if c.required else "") for c in spec.columns])
    ws.freeze_panes = "A2"
    for idx, col in enumerate(spec.columns, start=1):
        if col.enum is not None:
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(col.enum) + '"',
                allow_blank=not col.required,
            )
            letter = get_column_letter(idx)
            dv.add(f"{letter}2:{letter}1000")
            ws.add_data_validation(dv)

    d = wb.create_sheet("Dictionnaire")
    d.append(["Colonne", "Type", "Obligatoire", "Valeurs permises", "Alias acceptés", "Aide"])
    for col in spec.columns:
        d.append(
            [
                col.name,
                col.kind,
                "oui" if col.required else "non",
                ", ".join(col.enum) if col.enum else "",
                ", ".join(col.aliases) if col.aliases else "",
                col.help,
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_export(spec: EntitySpec, rows: list[dict[str, Any]]) -> bytes:
    """Exporte les données existantes au même format que le template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = spec.label[:31]
    names = [c.name for c in spec.columns]
    ws.append(names)
    for row in rows:
        ws.append([row.get(n) for n in names])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _parse_ws(ws: Any) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).rstrip("*").strip() if h is not None else "" for h in rows[0]]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in raw):
            continue
        out.append({headers[i]: raw[i] for i in range(len(headers)) if headers[i]})
    return out


def parse_sheet(content: bytes, sheet_title: str | None = None) -> list[dict[str, Any]]:
    """Lit la première feuille (ou `sheet_title`) → liste de dicts par en-tête."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    if sheet_title is not None and sheet_title in wb.sheetnames:
        ws = wb[sheet_title]
    else:
        ws = wb.worksheets[0]
    return _parse_ws(ws)


# --------------------------------------------------------------------- pôles (multi-feuilles)


@dataclass(frozen=True)
class PoleSpec:
    pole: str
    label: str
    entities: tuple[EntitySpec, ...]


def _add_data_sheet(wb: Any, spec: EntitySpec) -> None:
    ws = wb.create_sheet(spec.label[:31])
    ws.append([c.name + ("*" if c.required else "") for c in spec.columns])
    ws.freeze_panes = "A2"
    for idx, col in enumerate(spec.columns, start=1):
        if col.enum is not None:
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(col.enum) + '"',
                allow_blank=not col.required,
            )
            letter = get_column_letter(idx)
            dv.add(f"{letter}2:{letter}1000")
            ws.add_data_validation(dv)


def build_pole_template(pole: PoleSpec) -> bytes:
    """Classeur multi-feuilles : une feuille par entité + un dictionnaire global."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for spec in pole.entities:
        _add_data_sheet(wb, spec)
    d = wb.create_sheet("Dictionnaire")
    d.append(
        ["Entité", "Colonne", "Type", "Obligatoire", "Valeurs permises", "Alias acceptés", "Aide"]
    )
    for spec in pole.entities:
        for col in spec.columns:
            d.append(
                [
                    spec.label,
                    col.name,
                    col.kind,
                    "oui" if col.required else "non",
                    ", ".join(col.enum) if col.enum else "",
                    ", ".join(col.aliases) if col.aliases else "",
                    col.help,
                ]
            )
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def parse_pole(content: bytes, pole: PoleSpec) -> dict[str, list[dict[str, Any]]]:
    """Lit chaque feuille par entité → {entity: lignes}."""
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    out: dict[str, list[dict[str, Any]]] = {}
    for spec in pole.entities:
        title = spec.label[:31]
        out[spec.entity] = _parse_ws(wb[title]) if title in wb.sheetnames else []
    return out


def build_pole_export(pole: PoleSpec, data: dict[str, list[dict[str, Any]]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for spec in pole.entities:
        ws = wb.create_sheet(spec.label[:31])
        names = [c.name for c in spec.columns]
        ws.append(names)
        for row in data.get(spec.entity, []):
            ws.append([row.get(n) for n in names])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
